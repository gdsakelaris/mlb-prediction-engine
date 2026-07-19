"""Grade a predictions workbook against the actual box scores.

Re-colors the workbook IN PLACE once games are final, using the project's
own scraped per-player box scores (Data/mlb_game_batting.csv /
mlb_game_pitching.csv / mlb_games.csv) joined on the workbook's ID column
— PlayerId-exact, so no third-party box-score scraping or name matching.

One rule, three sheets: on Batter Props, Pitching Props and Games, every
gradeable cell whose event ACTUALLY OCCURRED is painted solid green
(#00B050); text stays regular black like every other cell, and
everything else keeps its served look. Every
graded cell answers the same literal question: DID THE STAT OCCUR.
Binary batter columns light up if the event happened (the HR cell if he
homered); O/U line columns light up if the OVER hit (K > 8.5 only if he
actually struck out 9+); the Winner cell (and its Win Prob) if the named
team won. Mean columns (xK, xTB, Away/Home Score, ...) have no yes/no
event and are left untouched.

On the Bets sheet, every WINNING bet has its whole row painted the same
green; losing / not-yet-settled rows keep the light board. Unders settle
too (the row carries its Side).

Re-running is safe and idempotent: pass 1 strips every green cell back
to the plain board (it also strips the retired reddish headline tint
from workbooks served before that tint was removed), then pass 2 grades
fresh.

Doubleheaders: prop rows carry a G# column, so every batter/pitcher row
grades against ITS OWN game's box line (the tag's G#-th final, schedule
order) — a game-1 prediction is never credited with a game-2 stat.
Games-sheet rows are matched per game: the tag's i-th row grades against
the day's i-th final for that matchup (schedule order); if only one of
the two games is final the tag's rows are skipped until both are in.
Bets rows carry no G#, so on a multi-final day they stay unsettled
rather than misgraded.

If some games are missing (they weren't final at the last scrape), their
rows are skipped and counted — run  python Scrapers/scrape_gamelogs.py
to pull the late finals, then grade again.

After painting, prints a backtest-style day report: per head,
n / actual vs stated rate / AUC / logloss / Brier against the day's base
rate — plus the over-50% pick ledger. One day is a small sample: treat
thin-head AUC as directional and use --all for the across-days
accumulation of the identical cell surface.

Usage:
    python Tools/4_grade_results.py                  # newest in Predictions/
    python Tools/4_grade_results.py path\\to\\file.xlsx
    python Tools/4_grade_results.py --all            # grade EVERY workbook

--all paints EVERY dated workbook in Predictions/ in place — the same
recolor as the single-file mode, idempotent on already-graded books —
then pools every day's cell surface into one cumulative backtest-style
report (same per-head table + pick ledger) plus the pooled Bets tally.
A workbook that can't be graded yet (today's slate before finals) is
skipped with a note; one that is open in Excel still counts in the
report but keeps its old paint.
"""
import argparse
import re
import sys
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

DATA_DIR = Path(__file__).resolve().parent.parent / "Data"
PRED_DIR = Path(__file__).resolve().parent.parent / "Predictions"

# the one grading color: solid green = the stat occurred
GREEN = "00B050"
HIT_FILL = PatternFill("solid", fgColor=GREEN)
NO_FILL = PatternFill(fill_type=None)
# reddish headline tints from retired board layouts — stripped on sight
RETIRED_TINTS = {"FBE3E9", "F5DBE2"}

# Bets sheet: a WINNING bet paints its whole row the same solid green; a
# losing or not-yet-settled row shows the light-green board it was
# served with.
BETS_WIN = PatternFill("solid", fgColor=GREEN)
BETS_BOARD = PatternFill("solid", fgColor="E7F3E2")

# batter columns -> did the event happen, from one box line
BAT_EVENTS = {
    "HR":          lambda s: s["HR"] >= 1,
    "Hit":         lambda s: s["H"] >= 1,
    "2+ Hits":     lambda s: s["H"] >= 2,
    "Single":      lambda s: (s["H"] - s["2B"] - s["3B"] - s["HR"]) >= 1,
    "Double":      lambda s: s["2B"] >= 1,
    "Triple":      lambda s: s["3B"] >= 1,
    "2+ TB":       lambda s: s["TB"] >= 2,
    "3+ TB":       lambda s: s["TB"] >= 3,
    "4+ TB":       lambda s: s["TB"] >= 4,
    "Run":         lambda s: s["R"] >= 1,
    "2+ Runs":     lambda s: s["R"] >= 2,
    "RBI":         lambda s: s["RBI"] >= 1,
    "2+ RBI":      lambda s: s["RBI"] >= 2,
    "H+R+RBI 2+":  lambda s: (s["H"] + s["R"] + s["RBI"]) >= 2,
    "H+R+RBI 3+":  lambda s: (s["H"] + s["R"] + s["RBI"]) >= 3,
    "H+R+RBI 4+":  lambda s: (s["H"] + s["R"] + s["RBI"]) >= 4,
    "BB":          lambda s: s["BB"] >= 1,
    "SB":          lambda s: s["SB"] >= 1,
    "K":           lambda s: s["SO"] >= 1,
    "2+ K":        lambda s: s["SO"] >= 2,
    "3+ K":        lambda s: s["SO"] >= 3,
}
BAT_SUM_COLS = ["PA", "AB", "R", "H", "2B", "3B", "HR", "RBI", "BB", "SO",
                "SB", "TB"]

# pitcher O/U column pattern -> the actual-stat key it grades
LINE_RE = re.compile(r"^(K|Outs|Hits|BB|ER) > (\d+(?:\.\d+)?)$")
LINE_STAT = {"K": "SO", "Outs": "outs", "Hits": "H", "BB": "BB", "ER": "ER"}
RUNS_RE = re.compile(r"^Runs > (\d+(?:\.\d+)?)$")
# per-team run lines on the Games sheet
TEAM_RUNS_RE = re.compile(r"^(Away|Home) Runs > (\d+(?:\.\d+)?)$")

# Bets 'pitcher <word> o<line>' -> the actual-stat key (prefix-matched
# after stripping 'pitcher '; no label is a prefix of another)
BET_PIT_STAT = {"strikeouts": "SO", "outs": "outs", "hits allowed": "H",
                "walks": "BB", "earned runs": "ER"}


def ip_to_outs(ip):
    """'5.2' -> 17 outs (MLB notation: .1/.2 = thirds)."""
    ip = float(ip)
    whole = int(ip)
    return whole * 3 + round((ip - whole) * 10)


def load_actuals(date):
    """(batters {pid: day-summed Series}, starters {pid: dict},
    games {'AWY@HOM': [dict, ...]}, batters_game {(pid, gamepk): Series},
    starters_game {(pid, gamepk): dict}) for one date, from the scraped
    logs. The games lists keep mlb_games.csv row order (the schedule's
    game order), so a doubleheader's game 1 is entry 0 and game 2 entry 1.

    The per-GAME dicts are the primary grading source — rows with a G#
    grade against their own game's box line. The day-summed dicts remain
    only as the fallback for rows without one."""
    gb = pd.read_csv(DATA_DIR / "mlb_game_batting.csv", encoding="utf-8-sig",
                     low_memory=False)
    gb = gb[gb["Date"] == date].copy()
    gb[BAT_SUM_COLS] = gb[BAT_SUM_COLS].apply(pd.to_numeric, errors="coerce")
    batters = {int(pid): grp[BAT_SUM_COLS].sum()
               for pid, grp in gb.groupby("PlayerId")}
    batters_game = {(int(pid), int(gpk)): grp[BAT_SUM_COLS].sum()
                    for (pid, gpk), grp in gb.groupby(["PlayerId", "GamePk"])}

    gp = pd.read_csv(DATA_DIR / "mlb_game_pitching.csv", encoding="utf-8-sig",
                     low_memory=False)
    gp = gp[(gp["Date"] == date) & (gp["GS"] == 1)]
    starters, starters_game = {}, {}
    for (pid, gpk), grp in gp.groupby(["PlayerId", "GamePk"]):
        r = grp.iloc[0]
        line = {"SO": float(r["SO"]), "H": float(r["H"]),
                "BB": float(r["BB"]), "ER": float(r["ER"]),
                "outs": ip_to_outs(r["IP"])}
        starters_game[(int(pid), int(gpk))] = line
        starters.setdefault(int(pid), line)   # first start = fallback value

    g = pd.read_csv(DATA_DIR / "mlb_games.csv", encoding="utf-8-sig")
    g = g[g["Date"] == date].dropna(subset=["AwayScore", "HomeScore"])
    games = {}
    for _, r in g.iterrows():
        a, h = float(r["AwayScore"]), float(r["HomeScore"])
        games.setdefault(f'{r["AwayTeam"]}@{r["HomeTeam"]}', []).append({
            "total": a + h, "away": a, "home": h,
            "gamepk": int(r["GamePk"]),
            "winner": r["HomeTeam"] if h > a else r["AwayTeam"]})
    return batters, starters, games, batters_game, starters_game


def _row_stats(per_game, day_dict, games, pid, tag, gnum):
    """Actual stats for one workbook row: with a Game tag and a G# the
    row grades against its OWN game's line (None until that game is
    final); without them fall back to the day sum. Returns None when
    unresolvable."""
    try:
        pid = int(pid)
    except (TypeError, ValueError):
        return None
    if tag is not None and gnum is not None:
        finals = games.get(tag, [])
        k = gnum - 1
        if k < 0 or k >= len(finals):
            return None                       # that game not final yet
        return per_game.get((pid, finals[k]["gamepk"]))
    return day_dict.get(pid)


def _ungrade(ws):
    """Strip every green graded cell back to the plain board (regular
    weight, default color), and drop any retired reddish headline tint,
    so grading is idempotent and old books repair cleanly."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            try:
                rgb = cell.fill.start_color.rgb
            except AttributeError:
                continue
            if not isinstance(rgb, str):
                continue
            tail = rgb[-6:].upper()
            if tail == GREEN:
                cell.fill = NO_FILL
                cell.font = Font()
            elif tail in RETIRED_TINTS:
                cell.fill = NO_FILL


def _mark(cell, occurred):
    """Paint one gradeable cell's fill green if its event occurred;
    leave it untouched otherwise. The font is never changed — _ungrade
    already reset any previously graded cell to the plain board, so
    text stays regular black everywhere."""
    if not occurred:
        return
    cell.fill = HIT_FILL


def _name_pid(ws):
    """{Name: pid} and {(Game, Name): pid} from a prop sheet, so a Bets
    row (which carries the player NAME, not the ID) can be settled off
    the same PlayerId-keyed actuals the prop grids use. (Game, Name)
    wins when present; plain Name is the single-game fallback."""
    hidx = {str(c.value): j for j, c in enumerate(ws[1], start=1)}
    byname, bygame = {}, {}
    if "Name" not in hidx or "ID" not in hidx:
        return byname, bygame
    gj = hidx.get("Game")
    for i in range(2, ws.max_row + 1):
        nm = ws.cell(row=i, column=hidx["Name"]).value
        pid = ws.cell(row=i, column=hidx["ID"]).value
        if nm is None or pid is None:
            continue
        byname[str(nm)] = int(pid)
        if gj is not None:
            bygame[(str(ws.cell(row=i, column=gj).value), str(nm))] = int(pid)
    return byname, bygame


def _bat_actual(s, prop):
    """The count a Bets batter-prop label prices, from one box line.
    Keyword order matters: 'H+R+RBI' before 'RBI', the specific hit
    types before 'hit'."""
    if "H+R+RBI" in prop:
        return s["H"] + s["R"] + s["RBI"]
    if "total bases" in prop:
        return s["TB"]
    if "HR" in prop:
        return s["HR"]
    if "RBI" in prop:
        return s["RBI"]
    if "single" in prop:
        return s["H"] - s["2B"] - s["3B"] - s["HR"]
    if "double" in prop:
        return s["2B"]
    if "hit" in prop:
        return s["H"]
    if "run" in prop:
        return s["R"]
    if "walk" in prop:
        return s["BB"]
    if "stolen base" in prop:
        return s["SB"]
    return None


def _settle_bet(row, batters, starters, games, bat_pid, pit_pid, bg, sg):
    """Did this Bets row win? True / False / None (can't settle yet: no
    final, unmatched player, or a doubleheader day we can't pin to one
    game — Bets rows carry no G#, so any multi-final matchup is
    unsettleable rather than misgraded against a day sum). `row` is
    {header: value}; `bat_pid` / `pit_pid` resolve a (game, name) to a
    PlayerId."""
    game, prop = str(row.get("Game", "")), str(row.get("Prop", ""))
    side, line = str(row.get("Side", "")), row.get("Line")

    def _line():
        try:
            return float(line)
        except (TypeError, ValueError):
            return None

    def _one_final():
        finals = games.get(game, [])
        return finals[0] if len(finals) == 1 else None

    if prop == "moneyline":                     # Side is the picked team
        f = _one_final()
        return f["winner"] == side if f else None
    if prop == "total runs":
        f, ln = _one_final(), _line()
        if f is None or ln is None:
            return None
        occ = f["total"] > ln
        return occ if side == "Over" else not occ
    if prop.startswith("pitcher "):             # "pitcher strikeouts o6.5"
        rest = prop[len("pitcher "):]
        for lbl, stat in BET_PIT_STAT.items():
            if rest.startswith(lbl):
                pid = pit_pid(game, row.get("Player"))
                f, ln = _one_final(), _line()
                if pid is None or f is None or ln is None:
                    return None
                a = (sg or {}).get((int(pid), f["gamepk"]))
                if a is None:
                    a = starters.get(int(pid))
                if a is None:
                    return None
                occ = a[stat] > ln
                return occ if side == "Over" else not occ
        return None
    pid = bat_pid(game, row.get("Player"))      # batter prop
    f, ln = _one_final(), _line()
    if pid is None or f is None or ln is None:
        return None
    s = (bg or {}).get((int(pid), f["gamepk"]))
    if s is None:
        s = batters.get(int(pid))
    if s is None:
        return None
    actual = _bat_actual(s, prop)
    if actual is None:
        return None
    occ = actual > ln
    return occ if side == "Over" else not occ


def _grade_bets(wb, batters, starters, games, stats, bg, sg):
    """Paint every WINNING bet row solid green; reset the rest to the
    light board first, so re-running is idempotent (the Bets sheet is
    skipped by _ungrade for exactly this reason). No-op rows without a
    Prop (the 'no captured odds' note) are left alone."""
    if "Bets" not in wb.sheetnames:
        return
    ws = wb["Bets"]
    headers = [str(c.value) for c in ws[1]]
    hidx = {h: j for j, h in enumerate(headers, start=1)}
    if not {"Game", "Prop", "Side"} <= set(hidx):
        return
    bp_name, bp_game = (_name_pid(wb["Batter Props"])
                        if "Batter Props" in wb.sheetnames else ({}, {}))
    pp_name, pp_game = (_name_pid(wb["Pitching Props"])
                        if "Pitching Props" in wb.sheetnames else ({}, {}))

    def bat_pid(g, nm):
        nm = None if nm is None else str(nm)
        return bp_game.get((g, nm)) or bp_name.get(nm)

    def pit_pid(g, nm):
        nm = None if nm is None else str(nm)
        return pp_game.get((g, nm)) or pp_name.get(nm)

    ncol = ws.max_column
    for i in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=i, column=hidx[h]).value for h in headers}
        if row.get("Prop") in (None, ""):
            continue                            # the "no captured odds" note
        # reset to board (undo any earlier win paint)
        for j in range(1, len(headers) + 1):
            c = ws.cell(row=i, column=j)
            c.fill = BETS_BOARD
            c.font = Font()
        won = _settle_bet(row, batters, starters, games, bat_pid, pit_pid,
                          bg, sg)
        stats["bets"] = stats.get("bets", 0) + 1
        if won:
            stats["bets_won"] = stats.get("bets_won", 0) + 1
            for j in range(1, ncol + 1):
                ws.cell(row=i, column=j).fill = BETS_WIN


class GradeError(RuntimeError):
    """A workbook that can't be graded (yet): no date in the filename, or
    no box scores for its date. Single-file mode exits on it; --all skips
    the workbook with a note and keeps going."""


def grade(path):
    m = re.match(r"(\d{4}-\d{2}-\d{2})", Path(path).stem)
    if not m:
        raise GradeError(f"can't read the game date from the filename: "
                         f"{path}")
    date = m.group(1)
    batters, starters, games, bg, sg = load_actuals(date)
    if not batters:
        raise GradeError(f"no box scores for {date} in "
                         f"Data/mlb_game_batting.csv — run  "
                         f"python Scrapers/scrape_gamelogs.py  first")

    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        if ws.title == "Bets":
            continue          # graded by _grade_bets (whole-row), not _ungrade
        _ungrade(ws)
    stats = {"cells": 0, "hit": 0, "missing_rows": 0}
    rows = []                 # (sheet, head, stated_p, occurred) per cell

    def headers_of(ws):
        return {str(c.value): j for j, c in enumerate(ws[1], start=1)}

    def _tag_gnum(ws, hidx, i):
        """(Game tag, G#) of a grid row, or (None, None) without them."""
        g_j, gn_j = hidx.get("Game"), hidx.get("G#")
        if g_j is None or gn_j is None:
            return None, None
        tag = str(ws.cell(row=i, column=g_j).value)
        try:
            return tag, int(ws.cell(row=i, column=gn_j).value)
        except (TypeError, ValueError):
            return None, None

    if "Batter Props" in wb.sheetnames:
        ws = wb["Batter Props"]
        hidx = headers_of(ws)
        cols = {h: j for h, j in hidx.items() if h in BAT_EVENTS}
        for i in range(2, ws.max_row + 1):
            pid = ws.cell(row=i, column=hidx["ID"]).value
            tag, gnum = _tag_gnum(ws, hidx, i)
            s = _row_stats(bg, batters, games, pid, tag, gnum)
            if s is None:
                stats["missing_rows"] += 1
                continue
            for h, j in cols.items():
                stats["cells"] += 1
                occ = bool(BAT_EVENTS[h](s))
                stats["hit"] += occ
                cell = ws.cell(row=i, column=j)
                if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                    rows.append(("Batter Props", h, float(cell.value), occ))
                _mark(cell, occ)

    if "Pitching Props" in wb.sheetnames:
        ws = wb["Pitching Props"]
        hidx = headers_of(ws)
        line_cols = [(h, j, LINE_RE.match(h)) for h, j in hidx.items()
                     if LINE_RE.match(h)]
        for i in range(2, ws.max_row + 1):
            pid = ws.cell(row=i, column=hidx["ID"]).value
            tag, gnum = _tag_gnum(ws, hidx, i)
            a = _row_stats(sg, starters, games, pid, tag, gnum)
            if a is None:
                stats["missing_rows"] += 1
                continue
            for h, j, mm in line_cols:
                stats["cells"] += 1
                occ = bool(a[LINE_STAT[mm.group(1)]] > float(mm.group(2)))
                stats["hit"] += occ
                cell = ws.cell(row=i, column=j)
                if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                    rows.append(("Pitching Props", h, float(cell.value), occ))
                _mark(cell, occ)

    if "Games" in wb.sheetnames:
        ws = wb["Games"]
        hidx = headers_of(ws)
        run_cols = [(h, j, float(RUNS_RE.match(h).group(1)))
                    for h, j in hidx.items() if RUNS_RE.match(h)]
        # team-total lines: 'Away Runs > 3.5' grades that side's score
        team_cols = [(h, j, TEAM_RUNS_RE.match(h).group(1).lower(),
                      float(TEAM_RUNS_RE.match(h).group(2)))
                     for h, j in hidx.items() if TEAM_RUNS_RE.match(h)]
        # Doubleheaders: the same "AWY@HOM" tag appears once per game, in
        # start-time order, and the finals list keeps the schedule's game
        # order — so match the sheet's i-th row for a tag to the i-th
        # final. If the finals don't yet cover every predicted game of
        # the tag (game 2 not final at the last scrape), skip the tag's
        # rows rather than grade two predictions against one game.
        need = Counter(str(ws.cell(row=i, column=hidx["Game"]).value)
                       for i in range(2, ws.max_row + 1))
        seen = Counter()
        for i in range(2, ws.max_row + 1):
            tag = str(ws.cell(row=i, column=hidx["Game"]).value)
            finals = games.get(tag, [])
            k = seen[tag]
            seen[tag] += 1
            g = finals[k] if len(finals) == need[tag] else None
            if g is None:
                stats["missing_rows"] += 1
                continue
            if "Winner" in hidx:
                cell = ws.cell(row=i, column=hidx["Winner"])
                stats["cells"] += 1
                occ = str(cell.value) == g["winner"]
                stats["hit"] += occ
                _mark(cell, occ)
                # Win Prob belongs to the named winner -> same outcome
                if "Win Prob" in hidx:
                    wp = ws.cell(row=i, column=hidx["Win Prob"])
                    if isinstance(wp.value, (int, float)) and 0 <= wp.value <= 1:
                        rows.append(("Games", "Winner", float(wp.value), occ))
                    _mark(wp, occ)
            for h, j, line in run_cols:
                stats["cells"] += 1
                occ = bool(g["total"] > line)
                stats["hit"] += occ
                cell = ws.cell(row=i, column=j)
                if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                    rows.append(("Games", h, float(cell.value), occ))
                _mark(cell, occ)
            for h, j, side, line in team_cols:
                stats["cells"] += 1
                occ = bool(g[side] > line)
                stats["hit"] += occ
                cell = ws.cell(row=i, column=j)
                if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                    rows.append(("Games", h, float(cell.value), occ))
                _mark(cell, occ)

    _grade_bets(wb, batters, starters, games, stats, bg, sg)

    # everything above is computed either way; a failed save only means the
    # paint didn't land (file open in Excel) — the stats/rows stay usable,
    # so --all can still pool the day into its cumulative report.
    try:
        wb.save(path)
        painted = True
    except PermissionError:
        painted = False
    return date, stats, rows, painted


def _rank_auc(p, y):
    """Mann-Whitney AUC (tie-aware, no sklearn); None if one class only."""
    y = np.asarray(y, dtype=bool)
    npos, nneg = int(y.sum()), int((~y).sum())
    if npos == 0 or nneg == 0:
        return None
    r = pd.Series(p, dtype=float).rank()
    return float((r[y].sum() - npos * (npos + 1) / 2) / (npos * nneg))


def day_report(rows, title="Day report: per-head, backtest-style"):
    """The backtest-style read of one graded day (or, via --all, of the
    whole accumulated record — same table, pooled rows).

    Per head (in board order, grouped by sheet): graded cells, the
    actual occurrence rate vs the stated average (gap = actual - stated),
    AUC, and logloss/Brier next to what a constant base-rate forecast
    scores. Then the over-50% pick ledger. Mean columns have no yes/no
    event and never enter. Single-day n is small — thin heads' AUC is
    directional."""
    if not rows:
        return
    eps = 1e-6
    by_head = {}
    for sheet, head, p, occ in rows:
        by_head.setdefault((sheet, head), []).append((p, occ))

    print(f"\n=== {title} ===")
    hdr = (f"  {'head':14s} {'n':>6s} {'actual%':>8s} {'stated%':>8s}"
           f" {'gap':>7s} {'AUC':>6s} {'logloss':>8s} {'base_ll':>8s}"
           f" {'brier':>7s} {'base_br':>8s}")
    # sheets in first-seen order; heads in first-seen order within each —
    # pooled workbooks with evolving boards still group cleanly
    for sheet in dict.fromkeys(s for s, _ in by_head):
        print(f"\n  -- {sheet} --")
        print(hdr)
        for (s2, head), pr in by_head.items():
            if s2 != sheet:
                continue
            p = np.clip(np.array([x for x, _ in pr], float), eps, 1 - eps)
            y = np.array([o for _, o in pr], float)
            rate, stated = float(y.mean()), float(p.mean())
            p0 = min(max(rate, eps), 1 - eps)
            ll = float(-(y * np.log(p) + (1 - y) * np.log(1 - p)).mean())
            ll0 = float(-(y * np.log(p0) + (1 - y) * np.log(1 - p0)).mean())
            br = float(((p - y) ** 2).mean())
            br0 = float(((p0 - y) ** 2).mean())
            auc = _rank_auc(p, y.astype(bool))
            print(f"  {head:14s} {len(y):6d} {rate:8.1%} {stated:8.1%}"
                  f" {rate - stated:+7.1%}"
                  + (f" {auc:6.3f}" if auc is not None else "      -")
                  + f" {ll:8.3f} {ll0:8.3f} {br:7.3f} {br0:8.3f}")

    over = [(h, p, o) for _, h, p, o in rows if p > 0.5]
    print("\n=== Over-50% picks ===")
    if not over:
        print("  none stated above 50% today.")
        return
    hit = sum(o for _, _, o in over)
    avg = sum(p for _, p, _ in over) / len(over)
    print(f"  overall: {len(over):,} picks -> {hit:,} hit "
          f"({hit / len(over):.1%}); stated avg {avg:.1%}")
    byh = {}
    for h, p, o in over:
        byh.setdefault(h, []).append((p, o))
    print(f"  {'head':14s} {'picks':>6s} {'hit':>5s} {'hit%':>7s}"
          f" {'stated':>7s} {'gap':>7s}")
    for h, b in sorted(byh.items(), key=lambda kv: -len(kv[1])):
        k = sum(o for _, o in b)
        r, a = k / len(b), sum(p for p, _ in b) / len(b)
        print(f"  {h:14s} {len(b):6d} {k:5d} {r:7.1%} {a:7.1%} {r - a:+7.1%}")


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("workbook", nargs="?", default=None,
                    help="predictions .xlsx (default: newest in Predictions/)")
    ap.add_argument("--all", action="store_true",
                    help="grade (paint) EVERY dated workbook in "
                         "Predictions/ in place, then pool all days into "
                         "one cumulative backtest-style report — incl. the "
                         "pooled Bets tally")
    args = ap.parse_args()

    if args.all:
        books = sorted(PRED_DIR.glob("[0-9]*.xlsx"))
        if not books:
            sys.exit(f"no workbooks in {PRED_DIR}")
        all_rows, days, tot, unpainted = [], 0, {}, []
        print("grading workbooks:")
        for p in books:
            try:
                _, s, rows, painted = grade(p)
            except GradeError as e:
                print(f"  ! {p.name}: {e}")
                continue
            days += 1
            all_rows += rows
            for k, v in s.items():
                tot[k] = tot.get(k, 0) + v
            line = f"  {p.name}: {s['cells']:,} cells, {s['hit']:,} occurred"
            if s.get("bets"):
                line += f"; bets won {s.get('bets_won', 0)}/{s['bets']}"
            if s.get("missing_rows"):
                line += f"; {s['missing_rows']} row(s) no box score"
            if not painted:
                unpainted.append(p.name)
                line += "  [NOT painted — open in Excel]"
            print(line)
        if not days:
            sys.exit("nothing gradeable yet — no dated workbook has box "
                     "scores.")
        print(f"\ngraded {days} workbook(s): {tot.get('cells', 0):,} cells "
              f"checked, {tot.get('hit', 0):,} stats occurred")
        if tot.get("bets"):
            print(f"Bets (all days): {tot.get('bets_won', 0)} of "
                  f"{tot['bets']} bet(s) won")
        if tot.get("missing_rows"):
            print(f"{tot['missing_rows']} row(s) had no final box score — "
                  f"run  python Scrapers/scrape_gamelogs.py  and re-run")
        if unpainted:
            print(f"not repainted (open in Excel): {', '.join(unpainted)} — "
                  f"still counted in the report")
        day_report(all_rows, title=f"Cumulative report: {days} day(s), "
                                   f"{len(all_rows):,} graded cells, "
                                   f"backtest-style")
        return

    path = args.workbook
    if path is None:
        books = sorted(PRED_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
        if not books:
            sys.exit(f"no workbooks in {PRED_DIR}")
        path = books[-1]
    try:
        date, s, rows, painted = grade(path)
    except GradeError as e:
        sys.exit(str(e))
    if not painted:
        sys.exit(f"{path} is open in Excel (it holds the file lock) — "
                 f"close it there, then run this again.")
    print(f"graded {path}")
    print("")
    print(f"  {date}: {s['cells']:,} cells checked, {s['hit']:,} stats "
          f"occurred (now solid green)")
    if s.get("bets"):
        print(f"  Bets sheet: {s.get('bets_won', 0)} of {s['bets']} bet(s) "
              f"won -> row highlighted solid green")
    if s["missing_rows"]:
        print(f"  {s['missing_rows']} row(s) had no final box score yet — "
              f"run  python Scrapers/scrape_gamelogs.py  and grade again")
    day_report(rows)


if __name__ == "__main__":
    main()
