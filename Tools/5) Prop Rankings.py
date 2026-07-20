"""Prediction-family quality rankings from the replay ledger.

Rewritten 2026-07-19 to mirror the old model's prop_rankings board;
re-scored same day as Score v5 — the composite now reflects THIS engine's
doctrine (probability quality vs base, out-of-sample calibration, CLV vs
the close) instead of the old project's top-10 pick board. Inputs:

  Model/artifacts/calib_rows.parquet          every graded (family, market,
                                              p, y) row from the full-season
                                              replay, RAW engine probabilities
                                              (p_cal is rebuilt HERE by
                                              day-block CROSS-FIT family
                                              Platt maps — the honest
                                              out-of-sample estimate of the
                                              served number; the live joblib
                                              is deliberately not applied to
                                              the rows it was fit on)
  Model/artifacts/gate_rows_cache.parquet     CLV gate rows (optional —
                                              GateN / VsClose stay blank
                                              until enough prices accrue)

One row per FAMILY (user decision 2026-07-19: no per-line sheet — "Hits >
x", never "Hits > 3.5" / "Hits > 4.5" separately). Metrics are computed
per priced line and AVERAGED across the family — the old board's own
family-row method, which keeps AUC/ECE honest (pooling lines would let
trivial cross-line ranking inflate them). Lift / Top10% / Base% read the
family's flagship line (its highest-base market, the first line it
prices), exactly as the old board did.

The diagnostic column set, the Score, the tiers, the sort, the console KEY
and the workbook styling are the old board's:

  Columns  LogLoss LLBase Edge% Brier BrierBase AUC Lift Top10% Base% ECE
           Slope Int MAE MAEBase MAEGain Bias MeanPred MeanAct Disp
           DispCal DispObs Acc  (display clusters by model importance —
           proper-scoring/edge, discrimination, calibration, error/level,
           dispersion, Acc last). Disp/DispCal/DispObs are kept for layout
           parity but stay blank: the ledger carries no expected-count
           means to price dispersion against.
  Score    Score v5 — the engine-doctrine composite on three pillars:
           probability skill (Edge% 40%: cross-fit calibrated log-loss
           gain vs base, counted ONCE), trustworthiness (AUC 20% +
           out-of-sample calibration 20%: ECE and slope on day-block
           cross-fit p), market edge (VsClose 10%: the family's CLV-gate
           log-loss edge vs the de-vigged close, entering only at the
           gate's own sample floor n >= 800 — below it the remaining
           weights renormalize to 100%, insufficient evidence is not a
           penalty), plus selection power (OR-lift 10%). Each metric a
           0-1 quality on FIXED anchors. MAE / Acc / Top10% / Brier /
           standalone LogLoss are DISPLAY-ONLY diagnostics: improper,
           degenerate, or duplicates of Edge%.
  Score_lo the day-block-bootstrap LOWER bound of Score (10th percentile
           over resampled days) — a thin family earns a wide CI.
           TIERS ARE CUT ON THIS COLUMN (the proven score); Score itself
           when the bootstrap is skipped.
  Tier     cut on Score_lo, on the FROZEN semantic anchors
           (1 ELITE ... 6 AVOID) — never re-fit to the run.
  Sort     TIER first, Score desc, then the v5 pillars as pre-declared
           tie-breakers.

Output: console table + Tools/PROP_RANKINGS.xlsx (Rankings + Legend),
navy-header serve-workbook styling.

Usage:
    python "Tools/5) Prop Rankings.py" [--boot 300] [--out FILE]
"""

import argparse
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
ART = ROOT / "Model" / "artifacts"
DEFAULT_OUT = ROOT / "Tools" / "PROP_RANKINGS.xlsx"
NAVY = "FF041E42"

# ---------------------------------------------------------------- families
# One row per ledger family, named for the combined line family it prices.
# Display names only — sorting is by Tier/Score.
FAMILY_NAMES = {
    "hr":   "Batter HR",
    "sb":   "Batter SB",
    "bb":   "Batter BB",
    "b1":   "Batter Single",
    "b2":   "Batter Double",
    "b3":   "Batter Triple",
    "h":    "Batter Hits > x",
    "r":    "Batter Runs > x",
    "rbi":  "Batter RBI > x",
    "bk":   "Batter K > x",
    "tb":   "Batter TB > x",
    "hrr":  "Batter H+R+RBI > x",
    "pk":   "Pitcher K > x",
    "pout": "Pitcher Outs > x",
    "pha":  "Pitcher Hits > x",
    "pbb":  "Pitcher BB > x",
    "per":  "Pitcher ER > x",
    "tot":  "Game Runs > x",
    "tt":   "Team Runs > x",
    "ml":   "Game Winner (Home ML)",
}
WINNER_FAMS = {"ml"}       # ~15-game slates: lift = top-1 confident, not top-10

# ---------------------------------------------------- score construction
# Score v5 pillar weights — this engine's doctrine, not the old board's
# pick-list frame: probability skill counted once, discrimination and
# out-of-sample calibration honesty weighted equally beside it, the CLV
# gate's market edge in the composite (light, per user decision
# 2026-07-19 — the gate itself remains the betting bar). When the gate
# sample is below its own floor, vsclose drops OUT and the rest
# renormalize — insufficient evidence is not a penalty.
V5_W = {"edge": 0.40, "auc": 0.20, "cal": 0.20, "lift": 0.10,
        "vsclose": 0.10}
GATE_MIN_N = 800          # the CLV gate's own sample floor
VSCLOSE_FULL = 0.02       # log-loss edge vs the close worth full quality

# day-block cross-fit of the family Platt map (honest out-of-sample p_cal)
CV_FOLDS, CV_MIN_TRAIN = 5, 300

# Day-block bootstrap: resample DAYS (multinomial day counts = per-day
# weights), recompute the whole composite per draw, keep the lower bound.
BOOT_B, SCORE_LCB_Q, BOOT_SEED = 300, 0.10, 20260719

# Frozen semantic anchors on the 0-100 composite, applied to Score_lo
# (the proven score) since v5 — a family's tier depends on ITSELF alone,
# never on how the rest of the board moved in a retrain. The v4 cut
# values were carried into v5 as its initial anchors (deliberate,
# 2026-07-19); re-anchoring stays a documented, versioned decision,
# never a fit to the run.
PROB_TIER_CUTS = ((72.0, "1 ELITE"), (60.0, "2 STRONG"), (48.0, "3 SOLID"),
                  (44.0, "4 DECENT"), (41.0, "5 LOW CEILING"),
                  (float("-inf"), "6 AVOID"))

# selection-power gates (odds-ratio lift) — the old blue-mark thresholds,
# used here only to word the Notes guidance
OR_TOP, OR_DEEP = 1.55, 2.5


def tier_of(score):
    if not np.isfinite(score):
        return "-"
    return next(t for c, t in PROB_TIER_CUTS if score >= c)


# ------------------------------------------------------------ diagnostics
def _ll(y, p):
    p = np.clip(p, 1e-6, 1 - 1e-6)
    return float(-np.mean(y * np.log(p) + (1 - y) * np.log(1 - p)))


def _auc(y, p):
    pos, neg = p[y == 1], p[y == 0]
    if not len(pos) or not len(neg):
        return np.nan
    r = pd.Series(p).rank().values
    return float((r[y == 1].sum() - len(pos) * (len(pos) + 1) / 2)
                 / (len(pos) * len(neg)))


def ece(p, y, bins=10):
    """Expected calibration error, equal-count bins."""
    q = pd.qcut(p, bins, duplicates="drop")
    df = pd.DataFrame({"p": p, "y": y, "q": q})
    g = df.groupby("q", observed=True)
    return float((g["p"].mean().sub(g["y"].mean()).abs()
                  * g.size().div(len(df))).sum())


def _slope_irls(z, y, w, a=0.0, b=1.0, iters=25):
    """1-D weighted logistic fit by Newton/IRLS -> (intercept, slope):
    reality refit on the served logit. Warm-startable for the bootstrap."""
    for _ in range(iters):
        mu = 1.0 / (1.0 + np.exp(-(a + b * z)))
        Wd = w * mu * (1 - mu)
        r = w * (y - mu)
        h00, h01, h11 = Wd.sum(), (Wd * z).sum(), (Wd * z * z).sum()
        det = h00 * h11 - h01 * h01
        if abs(det) < 1e-12:
            break
        g0, g1 = r.sum(), (r * z).sum()
        da, db = (h11 * g0 - h01 * g1) / det, (-h01 * g0 + h00 * g1) / det
        a, b = a + da, b + db
        if abs(da) + abs(db) < 1e-10:
            break
    return float(a), float(b)


def _crossfit_pcal(fam_df):
    """Day-block cross-fitted family Platt map: refit the 2-parameter
    logit-space map on each day fold's complement, calibrate the held-out
    fold — the honest estimate of how the SERVED (family-Platt) number
    behaves on days it hasn't seen. Applying the live joblib here would
    grade the calibrators on their own fit rows. Identity (raw p) on any
    fold where a fit is impossible — thin complement, single class,
    non-positive slope — mirroring fit_calibrators' identity fallback."""
    p = np.clip(fam_df["p"].to_numpy(dtype=float), 1e-4, 1 - 1e-4)
    y = fam_df["y"].to_numpy(dtype=float)
    z = np.log(p / (1 - p))
    day_id, _ = pd.factorize(fam_df["Date"], sort=True)
    fold = day_id % CV_FOLDS
    out = p.copy()
    for k in range(CV_FOLDS):
        te = fold == k
        tr = ~te
        if not te.any() or tr.sum() < CV_MIN_TRAIN or y[tr].min() == \
                y[tr].max():
            continue
        a, b = _slope_irls(z[tr], y[tr], np.ones(int(tr.sum())))
        if b <= 0:
            continue
        out[te] = 1.0 / (1.0 + np.exp(-(a + b * z[te])))
    return np.clip(out, 1e-4, 1 - 1e-4)


def top10_lift(df, p_col="p_cal", y_col="y"):
    """Daily top-10 hit rate over the base rate (how much the best picks
    beat blind betting)."""
    day = df.assign(d=pd.to_datetime(df["Date"]).dt.date)
    top = day.sort_values(p_col, ascending=False).groupby("d").head(10)
    base = df[y_col].mean()
    return float(top[y_col].mean() / base) if base > 0 else np.nan


def top1_lift(df, p_col="p_cal", y_col="y"):
    """Winner variant: the day's single most CONFIDENT game (top-10 makes
    no sense on a ~15-game slate) — picked-side hit rate over base."""
    y = df[y_col].to_numpy(dtype=float)
    p = df[p_col].to_numpy(dtype=float)
    base = y.mean()
    day = df.assign(d=pd.to_datetime(df["Date"]).dt.date,
                    conf=np.maximum(p, 1 - p),
                    hit=((p >= 0.5) == (y == 1)).astype(float))
    top1 = day.sort_values("conf", ascending=False).groupby("d").head(1)
    return float(top1["hit"].mean() / base) if base > 0 else np.nan


def or_lift(lift, base):
    """Odds-ratio top-pick lift: the top picks' hit ODDS over the base
    odds. Base-rate-fair where raw lift (capped at 1/base) is not, so a
    high-base family like Hit competes with a low-base one."""
    if not (np.isfinite(lift) and np.isfinite(base)) or not 0 < base < 1:
        return np.nan
    top = min(lift * base, 1 - 1e-6)
    return (top / (1 - top)) / (base / (1 - base))


def _binary_diags(p, y, lift):
    """The full ranked-diagnostic set for one priced line, from its served
    p / graded y rows."""
    base = float(y.mean())
    base_ll = _ll(y, np.full_like(p, base))
    ll = _ll(y, p)
    brier = float(np.mean((p - y) ** 2))
    brier_base = base * (1 - base)
    e = ece(p, y)
    intc, slope = _slope_irls(np.log(p / (1 - p)), y, np.ones_like(y))
    acc = float(np.mean((p >= 0.5) == (y > 0.5)))
    mae = float(np.mean(np.abs(p - y)))
    mae_base = 2 * base * (1 - base)     # MAE of the base-rate constant
    return {
        "rel": 100 * (base_ll - ll) / base_ll if base_ll > 0 else np.nan,
        "ll": ll,
        "ll_base": base_ll,
        "auc": _auc(y, p),
        "brier": brier,
        "brier_base": brier_base,
        "brier_rel": (100 * (brier_base - brier) / brier_base
                      if brier_base > 0 else np.nan),
        "ece": e,
        "ece_rel": e / (base * (1 - base)) if 0 < base < 1 else np.nan,
        "slope": slope,
        "int": intc,
        "acc": acc,
        "acc_base": max(base, 1 - base),
        "mae": mae,
        "mae_base": mae_base,
        "mae_gain": mae_base - mae,
        "mae_rel": (100 * (mae_base - mae) / mae_base
                    if mae_base > 0 else np.nan),
        "lift": lift,
        "top10": lift * base if np.isfinite(lift) else np.nan,
        "base": base,
        "mean_actual": base,
        "mean_pred": float(np.mean(p)),
    }


# the per-line diagnostics a family row averages (Lift/Top10%/Base% read
# the flagship line instead — averaging hit rates across bases is not fair)
AVG_KEYS = ("rel", "ll", "ll_base", "auc", "brier", "brier_base",
            "brier_rel", "ece", "ece_rel", "slope", "int", "acc",
            "acc_base", "mae", "mae_base", "mae_gain", "mae_rel",
            "mean_actual", "mean_pred")


# ------------------------------------------------------- composite score
# Each pillar maps to a 0-1 quality on a FIXED anchor; the Score is
# 100 * the weight-normalized sum over the pillars PRESENT:
#   edge_q     relative log-loss beat over base rate: 8% = full — the
#              engine's headline metric, counted exactly once
#   auc_q      (AUC - .5) / .20 (.70 = full)
#   cal_q      out-of-sample calibration honesty, mean of the ECE and
#              slope qualities (ECE scaled by base-rate variance, 10% of
#              Bernoulli scale = 0; slope full at 1.0, zero +/-0.5 away)
#              — both on day-block CROSS-FIT p, never in-sample
#   lift_q     odds-ratio top-pick lift: 1.0 -> 0, 2.5 -> full
#   vsclose_q  CLV-gate log-loss edge vs the de-vigged close: 0 at <= 0,
#              full at +0.02 — present only at GateN >= 800 (the gate's
#              floor); absent, the other pillars renormalize to 100%
def _clip(x):
    return float(max(0.0, min(1.0, x)))


def _q(v, fn):
    return fn(v) if np.isfinite(v) else 0.0


def _quals(m):
    orl = m["orl"] if "orl" in m else or_lift(m.get("lift", np.nan),
                                              m.get("base", np.nan))
    return {
        "edge": _q(m.get("rel", np.nan), lambda r: _clip(r / 8.0)),
        "auc": _q(m.get("auc", np.nan), lambda a: _clip((a - 0.5) / 0.20)),
        "cal": 0.5 * _q(m.get("ece_rel", np.nan),
                        lambda e: _clip(1 - e / 0.10))
        + 0.5 * _q(m.get("slope", np.nan),
                   lambda s: _clip(1 - abs(s - 1) / 0.5)),
        "lift": _q(orl, lambda o: _clip((o - 1) / 1.5)),
    }


def score_v5(m):
    q = _quals(m)
    w = dict(V5_W)
    vs = m.get("vsclose", np.nan)
    if np.isfinite(vs):
        q["vsclose"] = _clip(vs / VSCLOSE_FULL)
    else:
        w.pop("vsclose")
    return 100 * sum(w[k] * q[k] for k in w) / sum(w.values())


def play_note(m):
    """Data-driven usage guidance — calibration trust, selection power,
    shading direction — no hand judgment."""
    notes = []
    if m["ece"] <= 0.007:
        notes.append("calibrated -> price bets directly")
    elif m["ece"] >= 0.011:
        notes.append(f"probability level drifts (ECE {m['ece']:.3f}) -> "
                     "trust picks more than prices")
    orl, lift = m.get("orl", np.nan), m.get("lift", np.nan)
    if np.isfinite(orl):
        if orl >= OR_DEEP:
            notes.append(f"top picks {lift:.1f}x base (OR {orl:.1f}) -> "
                         "follow the list deep")
        elif orl >= OR_TOP:
            notes.append(f"top picks {lift:.1f}x (OR {orl:.1f}) -> "
                         "top 3-10 only")
        else:
            notes.append(f"picks {lift:.1f}x base (OR {orl:.1f}) -> "
                         "no selection power")
    slope = m.get("slope", np.nan)
    if np.isfinite(slope):
        if slope < 0.85:
            notes.append(f"overconfident (slope {slope:.2f}) -> shade "
                         "extreme probabilities toward the middle")
        elif slope > 1.15:
            notes.append(f"underconfident (slope {slope:.2f}) -> extremes "
                         "are even better than stated")
    return "; ".join(notes)


# ============================================================ bootstrap
# Weighted day-block bootstrap: a resample draws multinomial day COUNTS and
# every metric becomes an O(n) weighted mean / weighted rank-sum — nothing
# is re-sorted or re-priced per draw. The whole family composite is
# recomputed B times; Score_lo is the SCORE_LCB_Q percentile.
def _binent(b):
    b = np.clip(b, 1e-12, 1 - 1e-12)
    return -(b * np.log(b) + (1 - b) * np.log(1 - b))


def _prep_binlike(day_id, D, p, y):
    """Precompute everything one priced line needs so a resample is pure
    reweighting: per-day aggregates for edge/base/lift, fixed sort orders +
    tie blocks for weighted AUC/ECE, and a warm-start calibration line."""
    p = np.clip(np.asarray(p, float), 1e-4, 1 - 1e-4)
    y = np.asarray(y, float)
    z = np.log(p / (1 - p))
    nll = -(y * np.log(p) + (1 - y) * np.log(1 - p))
    sumy = np.bincount(day_id, y, D)
    cnt = np.bincount(day_id, minlength=D).astype(float)
    nllsum = np.bincount(day_id, nll, D)
    sqsum = np.bincount(day_id, (p - y) ** 2, D)          # Brier numerator
    aesum = np.bincount(day_id, np.abs(p - y), D)         # MAE numerator
    hitsum = np.bincount(day_id,                          # accuracy numerator
                         ((p >= 0.5) == (y > 0.5)).astype(float), D)
    # per-day top-10-by-p y-sum (fixed under day resampling)
    o = np.lexsort((-p, day_id))
    did_s, y_s = day_id[o], y[o]
    gstart = np.flatnonzero(np.concatenate(([True], did_s[1:] != did_s[:-1])))
    pos = np.arange(len(o)) - np.repeat(
        gstart, np.diff(np.append(gstart, len(o))))
    tm = pos < 10
    picks_sumy = np.bincount(did_s[tm], y_s[tm], D)
    picks_k = np.bincount(did_s[tm], minlength=D).astype(float)
    od = np.argsort(-p, kind="mergesort")           # AUC: descending, ties
    ps_d = p[od]
    starts = np.concatenate(([0], np.nonzero(np.diff(ps_d))[0] + 1))
    oa = np.argsort(p, kind="mergesort")            # ECE: ascending
    return {"D": D, "z": z, "y": y, "day_id": day_id,
            "sumy": sumy, "cnt": cnt, "nllsum": nllsum,
            "sqsum": sqsum, "aesum": aesum, "hitsum": hitsum,
            "picks_sumy": picks_sumy, "picks_k": picks_k,
            "ys_d": y[od], "did_d": day_id[od], "starts": starts,
            "pa": p[oa], "ya": y[oa], "did_a": day_id[oa],
            "ab0": _slope_irls(z, y, np.ones_like(y))}


def _wauc(pp, w):
    wr = w[pp["did_d"]]
    dTP = np.add.reduceat(wr * pp["ys_d"], pp["starts"])
    dFP = np.add.reduceat(wr * (1 - pp["ys_d"]), pp["starts"])
    P, N = dTP.sum(), dFP.sum()
    if P <= 0 or N <= 0:
        return np.nan
    return float(((np.cumsum(dTP) - dTP) * dFP
                  + 0.5 * dTP * dFP).sum() / (P * N))


def _wece(pp, w):
    wa = w[pp["did_a"]]
    cw = np.cumsum(wa)
    tot = cw[-1]
    if tot <= 0:
        return np.nan
    edges = np.linspace(0, tot, 11)[1:-1]
    bnd = np.unique(np.concatenate(([0], np.searchsorted(cw, edges),
                                    [len(wa)])))
    e = 0.0
    for a, b in zip(bnd[:-1], bnd[1:]):
        sw = wa[a:b].sum()
        if sw <= 0:
            continue
        mp = (wa[a:b] * pp["pa"][a:b]).sum() / sw
        my = (wa[a:b] * pp["ya"][a:b]).sum() / sw
        e += (sw / tot) * abs(mp - my)
    return float(e)


def _wslope(pp, w):
    return _slope_irls(pp["z"], pp["y"], w[pp["day_id"]],
                       a=pp["ab0"][0], b=pp["ab0"][1], iters=6)


def _wdiags(pp, Wm):
    """B ranked-diagnostic dicts for one priced line. Per-day quantities
    vectorize across all B draws; only AUC/ECE/cal-line loop."""
    Wtot = Wm @ pp["cnt"]
    base = (Wm @ pp["sumy"]) / Wtot
    base_ll = _binent(base)
    rel = np.where(base_ll > 0, 100 * (base_ll - (Wm @ pp["nllsum"]) / Wtot)
                   / base_ll, np.nan)
    brier = (Wm @ pp["sqsum"]) / Wtot
    bvar = base * (1 - base)
    brier_rel = np.where(bvar > 0, 100 * (bvar - brier) / bvar, np.nan)
    mae = (Wm @ pp["aesum"]) / Wtot
    mae_base = 2 * base * (1 - base)
    mae_rel = np.where(mae_base > 0, 100 * (mae_base - mae) / mae_base,
                       np.nan)
    acc = (Wm @ pp["hitsum"]) / Wtot
    pkw = Wm @ pp["picks_k"]
    lift = np.where((pkw > 0) & (base > 0),
                    (Wm @ pp["picks_sumy"]) / pkw / base, np.nan)
    ms = []
    for b, w in enumerate(Wm):
        bb = base[b]
        er = _wece(pp, w) / (bb * (1 - bb)) if 0 < bb < 1 else np.nan
        a_, s_ = _wslope(pp, w)
        ms.append({"rel": rel[b], "auc": _wauc(pp, w), "ece_rel": er,
                   "slope": s_, "int": a_, "brier_rel": brier_rel[b],
                   "acc": acc[b], "acc_base": max(bb, 1 - bb),
                   "mae_rel": mae_rel[b],
                   "top10": (lift[b] * bb if np.isfinite(lift[b])
                             else np.nan),
                   "orl": or_lift(lift[b], bb)})
    return ms


def _family_boot(preps, Wm, vsclose=np.nan):
    """B family composite Scores: per-line diagnostics averaged (as the
    point estimate does), Lift from the flagship line (preps[0]).
    vsclose is held constant across draws — the gate rows are a separate
    sample the board's day bootstrap does not resample."""
    per = [_wdiags(pp, Wm) for pp in preps]
    keys = ("rel", "auc", "ece_rel", "slope", "int", "brier_rel",
            "acc", "acc_base", "mae_rel")
    sc = np.empty(len(Wm))
    for b in range(len(Wm)):
        ds = [pl[b] for pl in per]
        m = {k: float(np.nanmean([d[k] for d in ds])) for k in keys}
        m["orl"] = ds[0]["orl"]
        m["top10"] = ds[0]["top10"]
        m["vsclose"] = vsclose
        sc[b] = score_v5(m)
    return sc


# ---------------------------------------------------------- provenance
def _provenance():
    """Ledger provenance line: replay window, row vintage, and whether
    the ENGINE artifacts have been retrained since the rows were made.
    The board silently outlives retrains — and during a frozen
    forward-test window the ledger CANNOT legally refresh (a
    --fit-calibrators rerun would rewrite the frozen calibrators) — so
    the board states what it reflects instead of implying live."""
    import datetime as dt
    rows_p = ART / "calib_rows.parquet"
    parts = []
    try:
        d = pd.read_parquet(rows_p, columns=["Date"])["Date"]
        parts.append(f"ledger {pd.Timestamp(d.min()).date()}.."
                     f"{pd.Timestamp(d.max()).date()} "
                     f"({pd.Series(d).nunique()} slates)")
    except Exception:                                   # noqa: BLE001
        parts.append("ledger window unreadable")
    rows_mt = rows_p.stat().st_mtime
    parts.append("replayed " + dt.datetime.fromtimestamp(rows_mt)
                 .strftime("%Y-%m-%d %H:%M"))
    # calibrators/heads are FIT FROM these rows (always newer — fine);
    # staleness = the ENGINE moving on without a re-replay
    newer = [f.split("_")[0] for f in
             ("a1_model.joblib", "a2_model.joblib",
              "hazard_model.joblib", "sb_models.joblib")
             if (ART / f).exists()
             and (ART / f).stat().st_mtime > rows_mt + 60]
    parts.append("engine retrained since ledger: "
                 + (", ".join(newer) if newer else "no"))
    return "; ".join(parts)


# ------------------------------------------------------------ build table
def _load_ledger():
    """calib_rows + day-block cross-fit family Platt p_cal — the honest
    out-of-sample view of the served calibration (the live joblib would
    be graded on its own fit rows here)."""
    rows_p = ART / "calib_rows.parquet"
    if not rows_p.exists():
        raise SystemExit("calib_rows.parquet missing — run a replay "
                         "(evaluate.py --fit-calibrators) first")
    df = pd.read_parquet(rows_p)
    parts = []
    for _, fam_df in df.groupby("family", sort=False):
        fam_df = fam_df.copy()
        fam_df["p_cal"] = _crossfit_pcal(fam_df)
        parts.append(fam_df)
    return pd.concat(parts, ignore_index=True)


def build_table(boot=BOOT_B, rows=None):
    df = rows if rows is not None else _load_ledger()
    # served = family Platt + residual head; the board grades the Platt
    # stage (cross-fit), so surface each active head's held-out gain as
    # display context — positive means serving beats the board's view
    heads_gain = {}
    hr_p = ART / "heads_report.csv"
    if hr_p.exists():
        try:
            hrep = pd.read_csv(hr_p)
            heads_gain = {
                str(r.family): float(r.gain) for r in hrep.itertuples()
                if getattr(r, "best_iter", 0) > 0
                and np.isfinite(getattr(r, "gain", np.nan))}
        except Exception:                               # noqa: BLE001
            pass

    gate = {}
    gp = ART / "gate_rows_cache.parquet"
    if gp.exists():
        g = pd.read_parquet(gp)
        for fam, sub in g.groupby("family"):
            y = sub["y"].astype(float).values
            if len(sub) >= 50 and 0 < y.mean() < 1:
                gate[fam] = dict(
                    n=len(sub),
                    vs_close=_ll(y, sub["p_close"].values)
                    - _ll(y, sub["p_model"].values))

    rng = np.random.default_rng(BOOT_SEED)
    rows = []
    for fam, name in FAMILY_NAMES.items():
        fam_df = df[df["family"] == fam].reset_index(drop=True)
        if fam_df.empty:
            continue
        day_id, uniq = pd.factorize(fam_df["Date"], sort=False)
        D = len(uniq)

        mkts = []                      # (market, base, diags, row mask)
        for mkt in fam_df["market"].unique():
            mask = (fam_df["market"] == mkt).to_numpy()
            sub = fam_df[mask]
            y = sub["y"].to_numpy(dtype=float)
            if len(sub) < 100 or y.min() == y.max():
                continue
            p = sub["p_cal"].to_numpy()
            lift = top1_lift(sub) if fam in WINNER_FAMS else top10_lift(sub)
            d = _binary_diags(p, y, lift)
            mkts.append((str(mkt), d["base"], d, mask))
        if not mkts:
            continue
        mkts.sort(key=lambda t: -t[1])       # flagship = highest base first

        ds = [d for _, _, d, _ in mkts]
        m = {k: float(np.nanmean([d[k] for d in ds])) for k in AVG_KEYS}
        fl = ds[0]
        m["lift"], m["top10"], m["base"] = fl["lift"], fl["top10"], fl["base"]
        m["orl"] = or_lift(m["lift"], m["base"])
        m["bias"] = m["mean_pred"] - m["mean_actual"]
        gv = gate.get(fam, {})
        m["vsclose"] = (gv["vs_close"]
                        if gv.get("n", 0) >= GATE_MIN_N else np.nan)
        score = score_v5(m)

        score_lo = np.nan
        if boot and D >= 5:
            Wm = rng.multinomial(D, np.full(D, 1.0 / D),
                                 size=boot).astype(float)
            preps = [_prep_binlike(day_id[mask], D,
                                   fam_df["p_cal"].to_numpy()[mask],
                                   fam_df["y"].to_numpy(dtype=float)[mask])
                     for _, _, _, mask in mkts]
            sc = _family_boot(preps, Wm, m["vsclose"])
            sc = sc[np.isfinite(sc)]
            if len(sc):
                score_lo = float(np.percentile(sc, SCORE_LCB_Q * 100))
        rows.append({
            "Market": name, "Key": fam, "Class": "Prob",
            "Score": score, "Score_lo": score_lo, "N": len(fam_df),
            "LogLoss": m["ll"], "LLBase": m["ll_base"], "Edge%": m["rel"],
            "Brier": m["brier"], "BrierBase": m["brier_base"],
            "AUC": m["auc"], "Lift": m["lift"],
            "Top10%": 100 * m["top10"], "Base%": 100 * m["base"],
            "ECE": m["ece"], "Slope": m["slope"], "Int": m["int"],
            "MAE": m["mae"], "MAEBase": m["mae_base"],
            "MAEGain": m["mae_gain"], "Bias": m["bias"],
            "MeanPred": m["mean_pred"], "MeanAct": m["mean_actual"],
            "Disp": np.nan, "DispCal": np.nan, "DispObs": np.nan,
            "Acc": m["acc"],
            "GateN": gv.get("n", np.nan),
            "VsClose": gv.get("vs_close", np.nan),
            "HeadGain": heads_gain.get(fam, np.nan),
            "Notes": play_note(m),
        })
    return _rank_and_tier(pd.DataFrame(rows))


# ------------------------------------------------------------ goal board

BATTER_FAMS = {"hr", "h", "tb", "r", "rbi", "bb", "sb",
               "b1", "b2", "b3", "hrr", "bk"}
PITCHER_FAMS = {"pk", "pout", "pha", "pbb", "per"}


def build_goal_board(rows):
    """The workbook-sort test, one row per market/column, plus the >50%
    reliability ladders — metric implementations live in
    Model/evaluate.py (shared with the --ab readout)."""
    import sys
    sys.path.insert(0, str(ROOT / "Model"))
    import evaluate as E
    led = rows[["Date", "market", "family", "p_cal", "y"]].rename(
        columns={"p_cal": "p"})
    board = E.goal_metrics(led)
    fam_of = rows.drop_duplicates("market").set_index("market")["family"]
    board.insert(1, "family", board["market"].map(fam_of))
    fam_rank = {f: i for i, f in enumerate(FAMILY_NAMES)}
    board = board.sort_values(
        ["family", "base"], ascending=[True, False],
        key=lambda s: s.map(fam_rank) if s.name == "family" else s)
    disp = pd.DataFrame({
        "Market": board["market"], "Family": board["family"],
        "N": board["n"], "Base%": (100 * board["base"]).round(1),
        "AUC": board["auc"].round(3),
        "Top10 Stated%": (100 * board["t10_stated"]).round(1),
        "Top10 Hit%": (100 * board["t10_hit"]).round(1),
        "Top10 Gap": (100 * board["t10_gap"]).round(1),
        "Trust Depth": board["trust_depth"],
        ">50% N": board["hi_n"],
        ">50% Stated%": (100 * board["hi_stated"]).round(1),
        ">50% Hit%": (100 * board["hi_hit"]).round(1),
        ">50% Gap": (100 * board["hi_gap"]).round(1),
    })
    bands = []
    for scope, fams in (("Batter", BATTER_FAMS),
                        ("Pitcher", PITCHER_FAMS),
                        ("Game", {"tot", "tt", "ml"})):
        b = E.reliability_bands(led[led["family"].isin(fams)])
        if len(b):
            b.insert(0, "Scope", scope)
            bands.append(b)
    if not bands:
        return disp, pd.DataFrame()
    bands = pd.concat(bands, ignore_index=True)
    bands = pd.DataFrame({
        "Scope": bands["Scope"], "Band": bands["band"],
        "N": bands["n"],
        "Stated%": (100 * bands["stated"]).round(1),
        "Hit%": (100 * bands["hit"]).round(1),
        "Gap": (100 * bands["gap"]).round(1)})
    return disp, bands


# ------------------------------------------------------------- sort order
# TIER first (cut on Score_lo — the proven score), Score desc, then the
# v5 pillars as pre-declared tie-breakers (display-only diagnostics
# last). NaNs sort last on every key.
RANK_SORT = (("Edge%", False), ("VsClose", False), ("AUC", False),
             ("_slope_dev", True), ("ECE", True), ("Lift", False),
             ("_int_dev", True), ("Brier", True), ("_bias_dev", True),
             ("MAE", True), ("LogLoss", True), ("Top10%", False),
             ("Acc", False))

# DISPLAY order (the old board's): most-to-least important for the model,
# in related clusters — proper-scoring/edge, discrimination, calibration,
# error/level, dispersion, Acc last. Every column present for every row
# even where a row leaves it empty.
_DIAG_ORDER = ["LogLoss", "LLBase", "Edge%", "Brier", "BrierBase",
               "AUC", "Lift", "Top10%", "Base%",
               "ECE", "Slope", "Int",
               "MAE", "MAEBase", "MAEGain", "Bias", "MeanPred",
               "MeanAct", "Disp", "DispCal", "DispObs", "Acc"]
COLS = (["#", "Market", "Key", "Class", "Tier", "Score", "Score_lo", "N"]
        + _DIAG_ORDER + ["GateN", "VsClose", "HeadGain", "Notes"])


def _rank_and_tier(df):
    d = df.copy()
    sc = pd.to_numeric(d["Score"], errors="coerce")
    lo = pd.to_numeric(d["Score_lo"], errors="coerce")
    d["Tier"] = [tier_of(s) for s in lo.where(lo.notna(), sc)]
    d["_t"] = d["Tier"].map(
        lambda t: int(t[0]) if str(t)[:1].isdigit() else 9)

    def _col(name):
        return (pd.to_numeric(d[name], errors="coerce") if name in d
                else pd.Series(np.nan, index=d.index))

    d["_s"] = _col("Score").fillna(-np.inf)
    d["_slope_dev"] = (_col("Slope") - 1).abs()
    d["_int_dev"] = _col("Int").abs()
    d["_bias_dev"] = _col("Bias").abs()
    keys, asc = ["_t", "_s"], [True, False]
    for i, (col, ascending) in enumerate(RANK_SORT):
        v = _col(col)
        d[f"_k{i}"] = v.fillna(np.inf if ascending else -np.inf)
        keys.append(f"_k{i}")
        asc.append(ascending)
    d = (d.sort_values(keys, ascending=asc)
           .drop(columns=[c for c in d.columns if c.startswith("_")])
           .reset_index(drop=True))
    d.insert(0, "#", range(1, len(d) + 1))
    return d[COLS]


# ----------------------------------------------------------------- legend
LEGEND = [
    ("Sort order", "TIER first (cut on Score_lo, the proven score), then "
     "Score descending, then the v5 pillars as pre-declared tie-breakers "
     "(Edge%, VsClose, AUC, slope deviation, ECE, Lift; display-only "
     "diagnostics after). COLUMN order is display grouping, not the "
     "Score ranking: proper-scoring/edge (LogLoss +LLBase, Edge%, Brier "
     "+BrierBase), then discrimination (AUC, Lift, Top10%, Base% "
     "context), then calibration (ECE, Slope, Int), then error/level "
     "(MAE +MAEBase/MAEGain, Bias, MeanPred/MeanAct), then dispersion, "
     "Acc last. Every column is present for every row even where a row "
     "leaves it empty."),
    ("Families", "ONE row per line family (user decision 2026-07-19: no "
     "per-line sheet) — 'Pitcher Hits > x' is every Hits line combined, "
     "never 'Hits > 3.5' alone. Metrics are computed per priced line "
     "through the day-block CROSS-FIT family Platt map — the honest "
     "out-of-sample estimate of the served number; applying the live "
     "calibrators here would grade them on their own fit rows — and "
     "averaged across the family, which keeps AUC/ECE honest (pooling "
     "lines would let trivial cross-line ranking inflate them). Lift / "
     "Top10% / Base% read the family's flagship line (its highest-base "
     "market)."),
    ("Score", "Score v5 (2026-07-19): the engine-doctrine composite on "
     "three pillars, each metric a 0-1 quality on FIXED anchors — "
     "probability skill (Edge% 40%, the base-relative log-loss gain, "
     "counted exactly once), trustworthiness (AUC 20%; out-of-sample "
     "calibration 20% = ECE and slope on cross-fit p, half each), "
     "market edge (VsClose 10%, the CLV-gate log-loss edge vs the "
     "de-vigged close, 0.02 = full, entering only at GateN >= 800 — "
     "below the floor the remaining weights renormalize to 100%: "
     "insufficient market evidence is not a penalty), plus selection "
     "power (odds-ratio Lift 10%). MAE / Acc / Top10% / Brier / "
     "standalone LogLoss are DISPLAY-ONLY: improper against 0/1 "
     "outcomes, degenerate on low-base props, or duplicates of Edge%. "
     "Single replay season — the two-year stability haircut returns "
     "when a second held-out year exists."),
    ("Score_lo", "The day-block-bootstrap LOWER bound of Score (10th "
     "percentile over resampled days) — the same bootstrap philosophy as "
     "the engine's paired accept bar. A thin family (SB, Triple-class "
     "events) earns a wide CI and a low Score_lo; read it as the family's "
     "PROVEN score. TIERS ARE CUT ON THIS COLUMN (on Score itself only "
     "when the bootstrap is skipped). VsClose is held constant across "
     "draws — the gate rows are a separate sample this bootstrap does "
     "not resample."),
    ("N", "Graded ledger rows across the family's lines (full-season "
     "replay)."),
    ("Tier", "1 ELITE / 2 STRONG: trust and act on these. 3 SOLID: usable "
     "with the noted caveat. 4 DECENT: a real, playable edge — shallower, "
     "but the picks are honest. 5 LOW CEILING: not a weak model — a low "
     "ceiling; the event is a handful of Bernoulli trials with a "
     "compressed true-p spread, so even a perfect forecaster could only "
     "separate players a little. 6 AVOID: don't act in isolation. Cut "
     "from SCORE_LO (the proven score) on FROZEN semantic anchors — a "
     "family's tier depends only on its own proven edge/calibration, "
     "never on how other families moved in a retrain."),
    ("Edge%", "IN SCORE: 40%, the dominant pillar. How much better the "
     "family prices the event than the base-rate guess (relative "
     "log-loss beat on cross-fit calibrated p), averaged across every "
     "line the family prices — the engine's headline metric, counted "
     "exactly once."),
    ("VsClose", "IN SCORE: 10% (with GateN, see below). The family's "
     "CLV-gate log-loss edge vs the de-vigged closing line — the "
     "engine's actual objective. 0 at no edge, full quality at +0.02. "
     "Enters the Score only at GateN >= 800 (the gate's own sample "
     "floor); below it the other pillars renormalize to 100%."),
    ("AUC", "IN SCORE: 20%. Ranking skill (0.5 = coin flip): can it put "
     "the players who DID do it above the ones who didn't? Computed per "
     "line, averaged across the family — pure within-line ranking, "
     "comparable across families. The skill ceiling calibration can't "
     "fake."),
    ("Slope", "IN SCORE: half of the 20% calibration pillar. Calibration "
     "slope: refit of reality on the CROSS-FIT logit (out-of-sample — "
     "never the calibrators graded on their own fit rows). 1.0 = "
     "probabilities move exactly as much as they should; below 1 = "
     "overconfident (shade extremes toward the middle), above 1 = "
     "underconfident (extremes better than stated)."),
    ("ECE", "IN SCORE: half of the 20% calibration pillar. Average gap "
     "between stated probability and reality across equal-count deciles "
     "on cross-fit p. 0 = perfect. Inside the Score it is scaled by the "
     "family's base-rate variance (an ECE of .006 is far worse on an 11% "
     "event than on a 60% one); the column shows the raw value."),
    ("Lift", "IN SCORE: 10%, in its base-rate-fair odds-ratio form. "
     "Daily top-10 hit rate over the base rate — the day's best picks, "
     "expressed as selection power (winner: the day's single most "
     "confident game). Measured on the family's flagship line. Feeds "
     "the Notes depth guidance."),
    ("Int", "DISPLAY-ONLY (was ranked in v4). Calibration intercept "
     "(log-odds): the systematic shift left after the slope. 0 = level; "
     "positive = events happen more often than stated across the "
     "board."),
    ("Brier", "DISPLAY-ONLY (was ranked in v4): its relative beat is a "
     "near-duplicate of Edge%. Mean squared error of the stated "
     "probability (lower better); BrierBase = the Brier a flat "
     "base-rate forecast would score."),
    ("Top10%", "DISPLAY-ONLY (was ranked in v4). The RAW daily top-10 "
     "hit rate — not base-rate-fair across markets (a high-base market "
     "posts a high Top10% with little skill); read it against Base%. "
     "Measured on the flagship line."),
    ("LogLoss", "DISPLAY-ONLY as a standalone number — only "
     "cross-market comparable through its base-relative form, which IS "
     "Edge%, the Score's dominant pillar. Its child LLBase = the log "
     "loss of always guessing the base rate."),
    ("MAE", "DISPLAY-ONLY (was ranked in v4): not a proper scoring rule "
     "against a 0/1 outcome — it rewards shading toward the majority "
     "class. MAEBase = the base-rate constant's MAE, MAEGain = MAEBase "
     "- MAE."),
    ("Acc", "DISPLAY-ONLY (was ranked in v4): degenerate on low-base "
     "props by construction (always-no wins). Plain accuracy of the "
     ">=50% call; acc of the always-majority pick is the honest "
     "reference."),
    ("Base% / MeanAct / MeanPred / Bias", "UNRANKED CONTEXT columns. "
     "Base% = how often the event actually happens — the blind-bet rate "
     "Top10% must be read against (the flagship line's rate). MeanAct = "
     "the mean actual outcome; MeanPred = the mean served probability; "
     "Bias = MeanPred - MeanAct, the family's level bias in probability "
     "points (positive = stated too high across the board)."),
    ("Disp / DispCal / DispObs", "Kept for layout parity with the old "
     "board but BLANK: they grade an expected-count column's observed "
     "dispersion against what its P(over) pricing assumes, and the "
     "current ledger carries no expected-count means. They fill in when "
     "the replay ledger starts logging count mu alongside the line "
     "probabilities."),
    ("GateN / VsClose", "CLV gate columns: prices graded so far and the "
     "family's logloss edge vs the de-vigged CLOSING line (positive = "
     "the engine beats the close). Shown from 50 graded prices; IN THE "
     "SCORE from GateN >= 800, the gate's own sample floor — below it "
     "the other pillars renormalize, so a rarely-scraped family is "
     "unproven vs the market, not punished. The only market-referenced "
     "numbers on the board — everything else is internal measurement."),
    ("HeadGain", "DISPLAY-ONLY. Held-out log-loss improvement of the "
     "family's residual head — the correction the serve path stacks ON "
     "TOP of the family Platt stage this board grades. Positive = the "
     "actually-served number is slightly better than the board states "
     "for this family; blank = no active head. Not in the Score: the "
     "heads are graded by the CLV gate on served probabilities."),
    ("Notes", "Data-driven usage guidance derived from the diagnostics — "
     "calibration trust (price bets directly vs trust picks more than "
     "prices), selection power (follow the list deep / top 3-10 only / "
     "no selection power, gated on the odds-ratio lift), and shading "
     "direction from the calibration slope."),
    ("Goal Board (sheet)", "The workbook-sort test, one row per served "
     "column: sort a column high-to-low — Top10 Stated%/Hit% = the mean "
     "stated probability and the actual hit rate of each slate's top-10 "
     "cells (positive Gap = the top of the sort beats its own stated "
     "number); the >50% columns are the same read on every cell above "
     "50%. Trust Depth = deepest per-slate top-d whose odds-ratio lift "
     "over the base rate keeps its slate-block-bootstrap 10th-percentile "
     "lower bound above 1.5 — how far down the sorted column selection "
     "power is PROVEN (0 = not even the top pick separates from a blind "
     "bet at this sample). All on day-block cross-fit calibrated p."),
    ("Reliability (sheet)", "Pooled stated-vs-hit ladder in 5-point "
     "bands above 50%, by scope (Batter / Pitcher / Game markets): the "
     "'cells above 50% should hit at their number, and more often the "
     "higher they go' check. Gap = Hit% - Stated%, in probability "
     "points (positive = reality beat the stated number)."),
    ("Data provenance", "Filled at write time — the ledger window this "
     "board was built from, when it was replayed, and whether the "
     "engine has retrained since (during a frozen forward-test window "
     "the ledger deliberately does not refresh; the board reflects the "
     "stack as of its replay date)."),
]


# ------------------------------------------------------------------ excel
_ND = [("Score", 0), ("Score_lo", 0), ("N", 0),
       ("Base%", 1), ("Lift", 2), ("Edge%", 2), ("Slope", 2),
       ("Int", 2), ("Disp", 2), ("DispCal", 2), ("DispObs", 2),
       ("Brier", 4), ("BrierBase", 4), ("ECE", 4), ("Bias", 2),
       ("MAE", 3), ("MAEBase", 3), ("MAEGain", 3), ("AUC", 3),
       ("Top10%", 1), ("LogLoss", 4), ("LLBase", 4), ("MeanAct", 2),
       ("MeanPred", 2), ("Acc", 3), ("GateN", 0), ("VsClose", 4),
       ("HeadGain", 4)]


def save_excel(df, path, provenance=None, goal=None, bands=None):
    """Rankings + Goal Board + Reliability + Legend, serve-workbook
    styling (navy header, centered, thin borders, frozen header,
    filter arrows)."""
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                 Side)
    from openpyxl.utils import get_column_letter

    xl = df.copy()
    for c, nd in _ND:
        if c in xl:
            xl[c] = xl[c].round(nd)
    legend = [(t, provenance if t == "Data provenance" and provenance
               else m) for t, m in LEGEND]
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        xl.to_excel(xw, sheet_name="Rankings", index=False)
        if goal is not None and len(goal):
            goal.to_excel(xw, sheet_name="Goal Board", index=False)
        if bands is not None and len(bands):
            bands.to_excel(xw, sheet_name="Reliability", index=False)
        pd.DataFrame(legend, columns=["Term", "Meaning"]).to_excel(
            xw, sheet_name="Legend", index=False)

    wb = openpyxl.load_workbook(path)
    thin = Side(style="thin", color="FFB7B7B7")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_nowrap = Alignment(horizontal="left", vertical="center")
    head_fill = PatternFill("solid", fgColor=NAVY)
    white_bold = Font(color="FFFFFFFF", bold=True)

    def _grid(name, cols):
        if name not in wb.sheetnames:
            return
        ws = wb[name]
        widths = [len(str(c)) for c in cols]
        for j in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=j)
            cell.fill = head_fill
            cell.font = white_bold
            cell.border = box
            cell.alignment = center
        for i in range(2, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                cell = ws.cell(row=i, column=j)
                cell.border = box
                cell.alignment = (left_nowrap
                                  if cols[j - 1] == "Notes" else center)
                v = cell.value
                widths[j - 1] = max(widths[j - 1],
                                    len("" if v is None else str(v)))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (f"A1:{get_column_letter(ws.max_column)}"
                              f"{ws.max_row}")
        for j in range(1, ws.max_column + 1):
            cap = 80 if cols[j - 1] == "Notes" else 22
            ws.column_dimensions[get_column_letter(j)].width = \
                min(widths[j - 1] + 4, cap)

    _grid("Rankings", COLS)
    if goal is not None and len(goal):
        _grid("Goal Board", list(goal.columns))
    if bands is not None and len(bands):
        _grid("Reliability", list(bands.columns))

    lg = wb["Legend"]
    for j in (1, 2):
        cell = lg.cell(row=1, column=j)
        cell.fill = head_fill
        cell.font = white_bold
        cell.border = box
        cell.alignment = center
    for i in range(2, lg.max_row + 1):
        lg.cell(row=i, column=1).font = Font(bold=True)
        lg.cell(row=i, column=1).alignment = Alignment(
            horizontal="left", vertical="top")
        lg.cell(row=i, column=2).alignment = left
    lg.column_dimensions["A"].width = 26
    lg.column_dimensions["B"].width = 110
    wb.save(path)


# ---------------------------------------------------------------- console
_FMT = {"Score": ".0f", "Score_lo": ".0f", "N": ".0f",
        "Base%": ".1f", "Lift": ".2f", "Edge%": ".2f", "Slope": ".2f",
        "Int": "+.2f", "Disp": ".2f", "DispCal": ".2f",
        "DispObs": ".2f", "Brier": ".4f", "BrierBase": ".4f",
        "ECE": ".4f", "Bias": "+.3f", "MAE": ".3f", "MAEBase": ".3f",
        "MAEGain": ".3f", "AUC": ".3f", "Top10%": ".1f",
        "LogLoss": ".4f", "LLBase": ".4f", "MeanAct": ".2f",
        "MeanPred": ".2f", "Acc": ".3f", "GateN": ".0f",
        "VsClose": ".4f", "HeadGain": "+.4f"}


def _fmt(frame):
    o = frame.copy()
    for c, f in _FMT.items():
        if c in o:
            o[c] = o[c].map(
                lambda v, f=f: format(v, f) if pd.notna(v) else "-")
    return o


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--boot", type=int, default=BOOT_B,
                    help="day-block bootstrap draws for Score_lo "
                         "(0 = skip)")
    ap.add_argument("--out", metavar="FILE", default=str(DEFAULT_OUT),
                    help="output Excel file "
                         "(default: Tools/PROP_RANKINGS.xlsx)")
    ap.add_argument("--verbose", "-v", action="store_true",
                    help="print the column KEY and the full rankings "
                         "table (default: a short summary — the full "
                         "board lives in the workbook)")
    args = ap.parse_args()

    rows = _load_ledger()
    df = build_table(boot=args.boot, rows=rows)
    goal, bands = build_goal_board(rows)

    if args.verbose:
        print("\n=== Prediction-family quality rankings — full-season "
              "replay ledger, day-block cross-fit (out-of-sample) "
              "calibrated probabilities ===\n")
        print("  KEY (columns in display clusters; Score v5 pillar "
              "weights in brackets):")
        print("  Tier     trust band, cut on SCORE_LO (the proven "
              "score): 1 ELITE bet it ... 6 AVOID don't")
        print("  Score    v5 pillar composite, 0-100: Edge% 40 | AUC 20 "
              "| calibration (ECE+Slope) 20 | VsClose 10 | Lift 10; "
              "weights renormalize when the gate sample < 800 | "
              "Score_lo its day-block-bootstrap lower bound — THE TIER "
              "CUT")
        print("  N        graded ledger rows across the family's lines")
        print("  LogLoss  [display] standalone log loss, family average "
              "(LLBase = the base-rate guess's — their relative gap IS "
              "Edge%)")
        print("  Edge%    [40%] relative log-loss beat over always "
              "guessing the base rate, averaged over every line priced")
        print("  Brier    [display] mean squared error of the stated "
              "probability (BrierBase = the base-rate forecast's)")
        print("  AUC      [20%] ranking skill per line, family average "
              "(0.5 = coin flip)")
        print("  Lift     [10%] daily top-10 hit rate / base rate on "
              "the flagship line (winner: top-1 confident); scored in "
              "its base-rate-fair odds-ratio form")
        print("  Top10%   [display] the RAW daily top-10 hit rate — "
              "base-inflated across markets; read against Base%")
        print("  Base%    context: how often the event actually happens "
              "(flagship line) — the blind-bet rate Top10% reads "
              "against")
        print("  ECE      [10%] average |stated% - actual%| across "
              "deciles on CROSS-FIT p (0 = perfectly calibrated "
              "out-of-sample)")
        print("  Slope    [10%] calibration slope on CROSS-FIT p: 1.00 "
              "= stated probabilities move exactly as much as reality "
              "(<1 overconfident)")
        print("  Int      [display] calibration intercept (log-odds): "
              "0 = level overall; positive = events beat their stated "
              "rate")
        print("  MAE      [display] mean |p - outcome| — improper "
              "against 0/1 outcomes, shown for continuity")
        print("  Bias     context: MeanPred - MeanAct, the family's "
              "level bias in probability points")
        print("  Disp     blank until the ledger logs count means "
              "(needs a mu column to price dispersion against)")
        print("  Acc      [display] accuracy of the >=50% call (sits at "
              "the base rate on low-base props by construction)")
        print("  GateN/VsClose  CLV gate: prices graded, logloss edge "
              "vs the de-vigged close [10% of Score at GateN >= 800; "
              "blank = too few prices yet]")
        print()
        with pd.option_context("display.width", 260,
                               "display.max_rows", 200,
                               "display.max_columns", 40):
            print(_fmt(df).to_string(index=False))
        print("\n  SORT: TIER first — cut on SCORE_LO, the proven score "
              "— then Score desc, then the v5 pillars as tie-breakers.")
        print("  One row per line FAMILY; per-line metrics are averaged "
              "across the family (Lift/Top10%/Base% from the flagship "
              "line). Calibration is day-block CROSS-FIT, never "
              "in-sample. Single replay season — no two-year stability "
              "haircut yet.")
        print()

    if args.verbose and len(goal):
        print("\n=== Goal Board — the workbook-sort test per served "
              "column (cross-fit calibrated p; Trust Depth = proven "
              "sorted-list depth, bootstrap-LCB odds-ratio gate) ===\n")
        with pd.option_context("display.width", 220,
                               "display.max_rows", 200):
            print(goal.to_string(index=False))

    prov = _provenance()
    save_excel(df, Path(args.out), provenance=prov, goal=goal,
               bands=bands)
    counts = df["Tier"].value_counts().to_dict()
    tiers = " | ".join(f"{t} {counts.get(t, 0)}"
                       for _, t in PROB_TIER_CUTS if counts.get(t, 0))
    print(f"Prop rankings (Score v5): {len(df)} families scored, "
          f"{int(df['N'].sum()):,} graded rows")
    print(f"  tiers: {tiers}")
    bb = bands[bands["Scope"] == "Batter"] if len(bands) else bands
    if len(bb):
        print("  >50% reliability, batter pooled (stated -> hit):")
        for _, r in bb.iterrows():
            print(f"    {r['Band']} n={int(r['N']):6d} "
                  f"{r['Stated%']:5.1f}% -> {r['Hit%']:5.1f}%  "
                  f"gap {r['Gap']:+.1f}")
    print(f"  {prov}")
    print(f"  written to {args.out}  (Rankings | Goal Board | "
          f"Reliability | Legend; -v for the full tables)")


if __name__ == "__main__":
    main()
