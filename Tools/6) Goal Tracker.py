"""Served goal-metric tracker: top-of-sort precision from the workbooks
the model ACTUALLY served, market by market, slate by slate.

For every dated workbook in Predictions/ (newest serve per date, same
rule as the grader), ranks each prop column by the served probability
and logs how the top of the sort did against the box scores:

  t1/t3/t5/t10_hit   hit rate of the top-K served rows that slate
  t10_stated         mean served p of the top 10 (reliability anchor)
  margin12           #1 -> #2 log-odds margin (top-slot separation,
                     the clear-leader signal, logged for calibration)
  scope              "all" rows, and "confirmed" = rows from games
                     whose BOTH lineups were confirmed at serve time
                     (away/home_lineup_src == "mlb" in the slate
                     archive; regen slates carry actual lineups and
                     count as confirmed)

Appends/upserts one row per (Date, market, scope) into
Model/artifacts/served_goal_tracker.csv (atomic write; re-runs replace
that date's rows, so tracking after late finals is safe). This is the
served-side series the W4.21 ranking-overlay decision reads — the
replay goal board measures the same thing offline; this file is the
live confirmation.

Usage:
    python "Tools/6) Goal Tracker.py"          # newest slate date
    python "Tools/6) Goal Tracker.py" --all    # backfill every date
"""
import argparse
import importlib
import json
import math
import os
import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
PRED_DIR = ROOT / "Predictions"
SLATE_DIR = ROOT / "Data" / "slates"
OUT_CSV = ROOT / "Model" / "artifacts" / "served_goal_tracker.csv"

sys.path.insert(0, str(TOOLS))
G4 = importlib.import_module("4) Grade Results")   # grading machinery

TOP_KS = (1, 3, 5, 10)
COLS = ["Date", "sheet", "market", "scope", "n", "n_games", "base",
        "t1_hit", "t3_hit", "t5_hit", "t10_hit", "t10_stated",
        "margin12", "n_nofinal"]


def newest_books():
    """{date: newest workbook Path} — repeat serves suffix _2, _3; the
    highest suffix is what was served last (grader convention)."""
    best = {}
    for p in sorted(PRED_DIR.glob("*.xlsx")):
        m = re.match(r"^(\d{4}-\d{2}-\d{2})(?:_(\d+))?$", p.stem)
        if not m:
            continue
        d, k = m.group(1), int(m.group(2) or 1)
        if d not in best or k > best[d][0]:
            best[d] = (k, p)
    return {d: p for d, (k, p) in best.items()}


def confirmed_tags(date):
    """Game tags ('AWY@HOM') whose both lineups were confirmed at serve
    time, from the slate archive. Regen slates (reconstructed actual
    lineups) count whole-slate confirmed. None = no slate found."""
    regen = SLATE_DIR / f"slate_{date}_regen.json"
    cands = ([regen] if regen.exists()
             else sorted(SLATE_DIR.glob(f"slate_{date}_*.json")))
    if not cands:
        return None
    slate = json.loads(cands[-1].read_text(encoding="utf-8"))
    games = slate["games"] if isinstance(slate, dict) else slate
    return {f'{g["away_team"]}@{g["home_team"]}' for g in games
            if g.get("away_lineup_src", "mlb") == "mlb"
            and g.get("home_lineup_src", "mlb") == "mlb"}


def _logit(p):
    p = min(max(float(p), 1e-6), 1 - 1e-6)
    return math.log(p / (1 - p))


def _metrics(rows):
    """One tracker record from [(p, y)] for a single market+slate."""
    rows = sorted(rows, key=lambda t: -t[0])
    ys = [y for _, y in rows]
    ps = [p for p, _ in rows]
    rec = dict(n=len(rows), base=round(sum(ys) / len(rows), 4))
    for k in TOP_KS:
        rec[f"t{k}_hit"] = round(sum(ys[:k]) / min(k, len(ys)), 4)
    rec["t10_stated"] = round(sum(ps[:10]) / min(10, len(ps)), 4)
    rec["margin12"] = (round(_logit(ps[0]) - _logit(ps[1]), 4)
                       if len(ps) >= 2 else None)
    return rec


def track_date(date, book):
    """Tracker records for one served workbook, or [] if no finals."""
    try:
        batters, starters, games, bg, sg = G4.load_actuals(date)
    except Exception as e:
        print(f"  ! {date}: box-score load failed ({e})")
        return []
    if not games:
        print(f"  - {date}: no finals yet, skipped")
        return []
    conf = confirmed_tags(date)
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    out = []
    for sheet, per_game, day_dict in (
            ("Batter Props", bg, batters), ("Pitching Props", sg, starters)):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c) for c in next(it)]
        idx = {c: j for j, c in enumerate(hdr)}
        if "ID" not in idx:
            continue
        markets = {}
        for c in hdr:
            if sheet == "Batter Props" and c in G4.BAT_EVENTS:
                markets[c] = ("event", G4.BAT_EVENTS[c])
            elif sheet == "Pitching Props":
                m = G4.LINE_RE.match(c)
                if m:
                    markets[c] = ("line", (G4.LINE_STAT[m.group(1)],
                                           float(m.group(2))))
        rowdata, nofinal = [], 0
        for r in it:
            tag = r[idx["Game"]] if "Game" in idx else None
            gnum = r[idx["G#"]] if "G#" in idx else None
            gnum = int(gnum) if isinstance(gnum, (int, float)) else None
            s = G4._row_stats(per_game, day_dict, games,
                              r[idx["ID"]], tag, gnum)
            if s is None:
                nofinal += 1
                continue
            rowdata.append((tag, r, s))
        for mkt, (kind, spec) in markets.items():
            j = idx[mkt]
            for scope in ("all", "confirmed"):
                rows, tags = [], set()
                for tag, r, s in rowdata:
                    if scope == "confirmed" and (conf is None
                                                 or tag not in conf):
                        continue
                    p = r[j]
                    if not isinstance(p, float):
                        continue
                    y = (int(bool(spec(s))) if kind == "event"
                         else int(float(s[spec[0]]) > spec[1]))
                    rows.append((p, y))
                    tags.add(tag)
                if not rows:
                    continue
                rec = _metrics(rows)
                rec.update(Date=date, sheet=sheet, market=mkt, scope=scope,
                           n_games=len(tags), n_nofinal=nofinal)
                out.append(rec)
    n_conf = "?" if conf is None else len(conf)
    print(f"  {date}: {len(out)} records "
          f"({n_conf} confirmed-lineup games in slate)")
    return out


def upsert(records):
    new = pd.DataFrame(records)[COLS]
    if OUT_CSV.exists():
        old = pd.read_csv(OUT_CSV, dtype={"Date": str})
        old = old[~old.Date.isin(set(new.Date))]
        new = pd.concat([old, new], ignore_index=True)
    new = new.sort_values(["Date", "sheet", "market", "scope"])
    fd, tmp = tempfile.mkstemp(dir=OUT_CSV.parent, suffix=".csv")
    os.close(fd)
    new.to_csv(tmp, index=False)
    os.replace(tmp, OUT_CSV)
    return new


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--all", action="store_true",
                    help="backfill every dated workbook")
    args = ap.parse_args()
    books = newest_books()
    dates = sorted(books) if args.all else [max(books)]
    print(f"tracking {len(dates)} slate date(s):")
    records = []
    for d in dates:
        records.extend(track_date(d, books[d]))
    if not records:
        print("nothing tracked — the newest slate has no finals yet "
              "(grade after the games finish, then rerun). Past slates "
              "are already in the CSV; use --all to re-sweep them.")
        if OUT_CSV.exists():
            done = pd.read_csv(OUT_CSV, dtype={"Date": str})
            print(f"{OUT_CSV.name} currently holds "
                  f"{done.Date.nunique()} slate dates "
                  f"({done.Date.min()}..{done.Date.max()})")
        return
    df = upsert(records)
    print(f"\n{OUT_CSV.name}: {len(df)} rows, "
          f"{df.Date.nunique()} slate dates")
    conf = df[df.scope == "confirmed"]
    if len(conf):
        pool = (conf.groupby("market")
                .agg(slates=("Date", "nunique"), t1=("t1_hit", "mean"),
                     t3=("t3_hit", "mean"), t10=("t10_hit", "mean"),
                     stated10=("t10_stated", "mean"),
                     med_m12=("margin12", "median"))
                .round(4).sort_values("t1", ascending=False))
        print("\nconfirmed-lineup served goal board (pooled across "
              "slates; t-cols = mean top-K hit rate):")
        print(pool.to_string())


if __name__ == "__main__":
    main()
