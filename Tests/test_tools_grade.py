"""Characterization tests for Tools/4 Grade Results: the pure settling
helpers (ip_to_outs, _bat_actual, _settle_bet, _row_stats) and one small
end-to-end grade() on tmp_path CSVs + workbook. GamePk/G#/DH *selection*
inside _settle_bet is already covered by test_gamepk.py — here the focus
is the stat mapping, the win/loss/push/None matrix on single-final games,
and the day-sum fallback firing ONLY when the matchup has exactly one
final. Everything runs on synthetic dicts / tmp_path; the live Data/ and
Predictions/ are never touched.
"""
import openpyxl
import pandas as pd
import pytest

from conftest import load_tool

t4 = load_tool("4) Grade Results")

# one final: BOS 4 @ NYY 5 (total 9), gamepk 111
F1 = {"total": 9.0, "away": 4.0, "home": 5.0, "gamepk": 111,
      "winner": "NYY"}
F2 = {"total": 7.0, "away": 6.0, "home": 1.0, "gamepk": 222,
      "winner": "BOS"}
GAMES1 = {"BOS@NYY": [F1]}
GAMES2 = {"BOS@NYY": [F1, F2]}

# one batter box line (pid 100) and one starter line (pid 300)
BAT = {"H": 2, "R": 1, "RBI": 3, "2B": 1, "3B": 0, "HR": 1, "TB": 6,
       "BB": 1, "SB": 0}
PIT = {"SO": 7.0, "H": 5.0, "BB": 2.0, "ER": 3.0, "outs": 17}


def _settle(row, games=GAMES1, batters=None, starters=None,
            bg=None, sg=None, bat_pid=None, pit_pid=None):
    return t4._settle_bet(
        row, batters or {}, starters or {}, games,
        bat_pid or (lambda g, n: 100), pit_pid or (lambda g, n: 300),
        bg if bg is not None else {(100, 111): BAT},
        sg if sg is not None else {(300, 111): PIT})


# ------------------------------------------------------------ ip_to_outs

def test_ip_to_outs():
    # MLB notation: the decimal is THIRDS of an inning, not tenths
    assert t4.ip_to_outs("5.2") == 17
    assert t4.ip_to_outs("5.1") == 16
    assert t4.ip_to_outs("5.0") == 15
    assert t4.ip_to_outs("0.1") == 1
    assert t4.ip_to_outs("0.2") == 2
    assert t4.ip_to_outs("9.0") == 27
    assert t4.ip_to_outs("0") == 0
    assert t4.ip_to_outs(6) == 18          # numeric input works too


# ------------------------------------------------------------ _bat_actual

def test_bat_actual_keyword_order_and_stat_map():
    s = BAT
    # combo label must NOT fall into the RBI branch (6, not 3)
    assert t4._bat_actual(s, "H+R+RBI 2+") == 6
    # specific hit types resolve before the generic 'hit' keyword
    assert t4._bat_actual(s, "single hit o0.5") == 0   # H-2B-3B-HR, not H
    assert t4._bat_actual(s, "double hit o0.5") == 1   # 2B, not H
    # each keyword maps to its own stat
    assert t4._bat_actual(s, "total bases o1.5") == 6
    assert t4._bat_actual(s, "HR o0.5") == 1
    assert t4._bat_actual(s, "RBI o1.5") == 3
    assert t4._bat_actual(s, "single o0.5") == 0
    assert t4._bat_actual(s, "double o0.5") == 1
    assert t4._bat_actual(s, "hits o1.5") == 2
    assert t4._bat_actual(s, "runs o0.5") == 1
    assert t4._bat_actual(s, "walks o0.5") == 1
    assert t4._bat_actual(s, "stolen bases o0.5") == 0
    # unknown label -> None (unsettleable, never a guess)
    assert t4._bat_actual(s, "fouls o0.5") is None
    # PINNED LANDMINE: a spelled-out 'home runs' label hits the 'run'
    # branch (case-sensitive 'HR' check misses it) and prices R, not HR —
    # bets labels must say 'HR'
    assert t4._bat_actual(s, "home runs o0.5") == s["R"]


# ------------------------------------------------------------ _settle_bet

def test_settle_moneyline():
    row = {"Game": "BOS@NYY", "Prop": "moneyline", "Side": "NYY"}
    assert _settle(row) is True
    assert _settle({**row, "Side": "BOS"}) is False
    # no final for the matchup -> unsettleable
    assert _settle({**row, "Game": "SEA@TEX"}) is None


def test_settle_total_runs():
    def tot(side, line):
        return {"Game": "BOS@NYY", "Prop": "total runs", "Side": side,
                "Line": line}
    assert _settle(tot("Over", 8.5)) is True
    assert _settle(tot("Under", 8.5)) is False
    assert _settle(tot("Over", 9.5)) is False
    assert _settle(tot("Under", 9.5)) is True
    assert _settle(tot("Over", 9)) == "push"       # landed exactly on 9
    assert _settle(tot("Under", 9)) == "push"
    assert _settle(tot("Over", None)) is None      # missing line
    assert _settle(tot("Over", "abc")) is None     # unparseable line


def test_settle_all_five_pitcher_labels():
    # read the dict: five labels, none a prefix of another (the
    # startswith match depends on that)
    assert len(t4.BET_PIT_STAT) == 5
    labels = list(t4.BET_PIT_STAT)
    for a in labels:
        for b in labels:
            assert a == b or not b.startswith(a)
    for lbl, stat in t4.BET_PIT_STAT.items():
        v = PIT[stat]
        over = {"Game": "BOS@NYY", "Prop": f"pitcher {lbl} o{v - 0.5}",
                "Side": "Over", "Line": v - 0.5}
        assert _settle(over) is True, lbl
        assert _settle({**over, "Side": "Under"}) is False, lbl
        assert _settle({**over, "Line": v}) == "push", lbl
    # unknown pitcher label falls through the loop -> None
    assert _settle({"Game": "BOS@NYY", "Prop": "pitcher quality starts o0.5",
                    "Side": "Over", "Line": 0.5}) is None
    # unmatched pitcher name -> None
    assert _settle({"Game": "BOS@NYY", "Prop": "pitcher strikeouts o6.5",
                    "Side": "Over", "Line": 6.5},
                   pit_pid=lambda g, n: None) is None


def test_settle_batter_prop():
    row = {"Game": "BOS@NYY", "Prop": "hits o1.5", "Side": "Over",
           "Line": 1.5}
    assert _settle(row) is True                    # H=2 > 1.5
    assert _settle({**row, "Side": "Under"}) is False
    # integer line landed exactly -> push
    assert _settle({**row, "Line": 2}) == "push"
    # unmatched player / missing line / unknown stat keyword -> None
    assert _settle(row, bat_pid=lambda g, n: None) is None
    assert _settle({**row, "Line": None}) is None
    assert _settle({"Game": "BOS@NYY", "Prop": "fouls o0.5",
                    "Side": "Over", "Line": 0.5}) is None


def test_settle_day_sum_fallback_only_on_single_final_days():
    row = {"Game": "BOS@NYY", "Prop": "hits o1.5", "Side": "Over",
           "Line": 1.5}
    # no per-game box line, ONE final -> the day sum settles it
    assert _settle(row, bg={}, batters={100: BAT}) is True
    # two finals: same row pinned to game 1 via G#, still no per-game
    # line -> stays unsettled (a day sum would mix both games)
    assert _settle({**row, "G#": 1}, games=GAMES2, bg={},
                   batters={100: BAT}) is None
    # same rule for the pitcher path
    prow = {"Game": "BOS@NYY", "Prop": "pitcher strikeouts o6.5",
            "Side": "Over", "Line": 6.5}
    assert _settle(prow, sg={}, starters={300: PIT}) is True
    assert _settle({**prow, "G#": 1}, games=GAMES2, sg={},
                   starters={300: PIT}) is None


# ------------------------------------------------------------ _row_stats

def test_row_stats():
    per_game = {(100, 111): "g1box", (100, 222): "g2box"}
    day = {100: "daysum"}
    games = {"BOS@NYY": [{"gamepk": 111}, {"gamepk": 222}]}
    # G#-tagged rows grade against their OWN game's line
    assert t4._row_stats(per_game, day, games, 100, "BOS@NYY", 1) == "g1box"
    assert t4._row_stats(per_game, day, games, 100, "BOS@NYY", 2) == "g2box"
    # tagged but that game not final (or nonsense ordinal) -> None
    assert t4._row_stats(per_game, day, games, 100, "BOS@NYY", 3) is None
    assert t4._row_stats(per_game, day, games, 100, "BOS@NYY", 0) is None
    one = {"BOS@NYY": [{"gamepk": 111}]}
    assert t4._row_stats(per_game, day, one, 100, "BOS@NYY", 2) is None
    # untagged -> day-sum fallback; float pid coerces
    assert t4._row_stats(per_game, day, games, 100, None, None) == "daysum"
    assert t4._row_stats(per_game, day, games, 100.0, None, None) == "daysum"
    # unknown pid / unparseable pid -> None
    assert t4._row_stats(per_game, day, games, 999, None, None) is None
    assert t4._row_stats(per_game, day, games, "abc", None, None) is None
    assert t4._row_stats(per_game, day, games, None, None, None) is None


# ------------------------------------------------------------ grade() e2e

DATE = "2026-07-22"


def _write_actuals(tmp_path):
    pd.DataFrame([{
        "GamePk": 111, "Date": DATE, "PlayerId": 100, "PA": 4, "AB": 4,
        "R": 1, "H": 2, "2B": 1, "3B": 0, "HR": 1, "RBI": 2, "BB": 0,
        "SO": 1, "SB": 0, "TB": 6,
    }]).to_csv(tmp_path / "mlb_game_batting.csv", index=False)
    pd.DataFrame([{
        "GamePk": 111, "Date": DATE, "PlayerId": 300, "GS": 1, "IP": "5.2",
        "SO": 7, "H": 5, "BB": 2, "ER": 3,
    }]).to_csv(tmp_path / "mlb_game_pitching.csv", index=False)
    pd.DataFrame([{
        "GamePk": 111, "Date": DATE, "AwayTeam": "BOS", "HomeTeam": "NYY",
        "AwayScore": 4, "HomeScore": 5,
    }]).to_csv(tmp_path / "mlb_games.csv", index=False)


def _write_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batter Props"
    ws.append(["Game", "Team", "G#", "Slot", "Name", "ID", "HR", "Hit"])
    ws.append(["BOS@NYY", "BOS", 1, 1, "Slugger Sam", 100, 0.25, 0.60])
    ws.append(["BOS@NYY", "BOS", 1, 2, "Missing Mo", 999, 0.10, 0.40])
    pp = wb.create_sheet("Pitching Props")
    pp.append(["Game", "Team", "G#", "Name", "ID", "K > 6.5"])
    pp.append(["BOS@NYY", "NYY", 1, "Ace Al", 300, 0.70])
    gs = wb.create_sheet("Games")
    gs.append(["Game", "Winner", "Win Prob", "Runs > 8.5", "Away Runs > 4.5"])
    gs.append(["BOS@NYY", "NYY", 0.55, 0.50, 0.30])
    bs = wb.create_sheet("Bets")
    bs.append(["Game", "Player", "Prop", "Side", "Line", "GamePk", "G#"])
    bs.append(["BOS@NYY", None, "moneyline", "NYY", None, 111, 1])
    bs.append(["BOS@NYY", "Slugger Sam", "hits o1.5", "Over", 1.5, 111, 1])
    bs.append(["BOS@NYY", None, "total runs", "Over", 9, 111, 1])
    bs.append(["BOS@NYY", None, None, None, None, None, None])  # odds note
    wb.save(path)


EXPECT_STATS = {
    # Batter row 1: HR+Hit (both hit); pid-999 row has no box -> missing.
    # Pitching: K>6.5 (7 -> hit). Games: Winner ok, Runs>8.5 ok (9),
    # Away Runs>4.5 no (4). Bets: ML win + hits-over win + total push;
    # the blank-Prop note row is skipped; bets_open never increments so
    # the key is absent.
    "cells": 6, "hit": 5, "missing_rows": 1,
    "bets": 3, "bets_settled": 2, "bets_won": 2, "bets_push": 1,
}


def test_grade_end_to_end_and_idempotent(tmp_path, monkeypatch):
    monkeypatch.setattr(t4, "DATA_DIR", tmp_path)
    _write_actuals(tmp_path)
    path = tmp_path / f"{DATE} preds.xlsx"
    _write_workbook(path)

    date, stats, rows, painted = t4.grade(path)
    assert date == DATE
    assert painted is True
    assert stats == EXPECT_STATS
    # every probability-valued graded cell lands in the reliability rows
    assert len(rows) == 6
    assert ("Batter Props", "HR", 0.25, True) in rows
    assert ("Games", "Away Runs > 4.5", 0.30, False) in rows

    # the paint: hit cells solid green, misses untouched, winning bet
    # rows green, the push row back on the light board
    wb = openpyxl.load_workbook(path)

    def tail(cell):
        rgb = cell.fill.start_color.rgb
        return rgb[-6:].upper() if isinstance(rgb, str) else None

    bat = wb["Batter Props"]
    assert tail(bat.cell(row=2, column=7)) == t4.GREEN        # HR hit
    g = wb["Games"]
    assert tail(g.cell(row=2, column=2)) == t4.GREEN          # Winner
    assert tail(g.cell(row=2, column=5)) != t4.GREEN          # Away Runs miss
    bets = wb["Bets"]
    assert tail(bets.cell(row=2, column=1)) == t4.GREEN       # ML win
    assert tail(bets.cell(row=4, column=1)) == "E7F3E2"       # push -> board

    # grading again is a no-op on the numbers (ungrade+regrade)
    date2, stats2, rows2, painted2 = t4.grade(path)
    assert (date2, stats2, rows2, painted2) == (date, stats, rows, painted)


def test_grade_errors(tmp_path, monkeypatch):
    monkeypatch.setattr(t4, "DATA_DIR", tmp_path)
    _write_actuals(tmp_path)
    # no date in the filename -> GradeError before any I/O
    with pytest.raises(t4.GradeError):
        t4.grade(tmp_path / "nodate.xlsx")
    # a date with no box scores -> GradeError before the workbook loads
    with pytest.raises(t4.GradeError):
        t4.grade(tmp_path / "2026-01-01.xlsx")
