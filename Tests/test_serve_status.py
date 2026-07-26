"""Degraded-serve banner (predict._status_sheet): a missing-calibrator
serve must be unmissable — red sheet at index 0, opened active; a clean
serve must add nothing.

Degraded-list lifetimes (predict_slate / main): Predictor.degraded is
process-lifetime ARTIFACT health only; per-game serve failures are
scoped to the predict_slate call that produced them, so a failed
afternoon serve can't stamp a false FAILED banner on a later clean
serve (or on the next --json slate file) from the same Predictor.
"""
import datetime as dt
import json
import sys

import pytest
from openpyxl import Workbook

import predict as PR


def test_clean_serve_adds_no_sheet():
    wb = Workbook()
    wb.active.title = "Batter Props"
    PR._status_sheet(wb, [], "2026-07-24")
    assert wb.sheetnames == ["Batter Props"]


def test_degraded_serve_banner_first_and_active():
    wb = Workbook()
    wb.active.title = "Batter Props"
    wb.create_sheet("Bets")
    PR._status_sheet(wb, ["output_calibrators.joblib MISSING — serving "
                          "RAW probabilities"], "2026-07-24")
    assert wb.sheetnames[0] == "!! STATUS"
    assert wb.active.title == "!! STATUS"
    v = wb["!! STATUS"].cell(row=1, column=1).value
    assert "DEGRADED SERVE 2026-07-24" in v and "RAW probabilities" in v
    # existing sheets untouched
    assert "Batter Props" in wb.sheetnames and "Bets" in wb.sheetnames


# ---------------------------------------------- degraded-list scoping

def _spec(away, home):
    return {"away_team": away, "home_team": home,
            "date": dt.date.today().isoformat(),
            "away_lineup": [], "home_lineup": []}


def _prep(spec, n_sims):
    if spec["away_team"] == "BAD":
        raise ValueError("boom")
    return {}, {"season": 2026}


def _no_batch(specs, n_sims, tick):
    raise RuntimeError("device unavailable")


def _bare_predictor(monkeypatch, artifact_msgs):
    """Predictor with __init__ skipped: only what predict_slate reads.
    _slate_batch raises so the default SERVE_BATCH=1 route exercises
    the batch-exception fallback into the per-game loop — the exact
    path the audit finding flagged."""
    P = PR.Predictor.__new__(PR.Predictor)
    P.degraded = list(artifact_msgs)
    P.calib = {}
    P.mkt_blend = False          # pure serve: no market-fair lookup
    P._slate_batch = _no_batch
    P.prepare_game = _prep
    monkeypatch.setattr(
        PR.F, "slate_context",
        lambda specs: (_ for _ in ()).throw(RuntimeError("no ctx")))
    monkeypatch.setattr(PR.sim, "run", lambda prep, **kw: {})
    return P


def test_game_failure_scoped_to_single_serve(monkeypatch):
    P = _bare_predictor(monkeypatch, [])
    out1 = P.predict_slate([_spec("BAD", "NYY"), _spec("ATL", "NYM")],
                           n_sims=4)
    assert len(out1) == 1
    (fmsg,) = out1[0]["degraded"]
    assert "game BAD@NYY" in fmsg and "FAILED" in fmsg
    assert "slate served WITHOUT it" in fmsg
    # the failure must not outlive its own predict_slate call
    assert P.degraded == []
    out2 = P.predict_slate([_spec("BOS", "TBR"), _spec("ATL", "NYM")],
                           n_sims=4)
    assert len(out2) == 2
    assert all(r["degraded"] == [] for r in out2)
    # and the clean serve's workbook gets NO red banner from serve 1
    wb = Workbook()
    wb.active.title = "Batter Props"
    PR._status_sheet(wb, out2[0]["degraded"],
                     dt.date.today().isoformat())
    assert wb.sheetnames == ["Batter Props"]


def test_artifact_messages_persist_game_failures_do_not(monkeypatch):
    art = ("output_calibrators.joblib MISSING — serving RAW "
           "probabilities; fix: python Model/evaluate.py "
           "--fit-calibrators --reuse-rows")
    P = _bare_predictor(monkeypatch, [art])
    out1 = P.predict_slate([_spec("BAD", "NYY"), _spec("ATL", "NYM")],
                           n_sims=4)
    assert out1[0]["degraded"][0] == art
    assert "FAILED" in out1[0]["degraded"][1]
    out2 = P.predict_slate([_spec("BOS", "TBR")], n_sims=4)
    # artifact health persists; serve 1's game failure does not
    assert out2[0]["degraded"] == [art]
    assert P.degraded == [art]


def test_headless_exit3_from_slate_game_failure(monkeypatch, tmp_path):
    """main() over two --json slates: a game failure in slate a must
    reach exit 3 and slate a's workbook, but never slate b's."""
    P = _bare_predictor(monkeypatch, [])
    slates = []
    for name, teams in (("a.json", [("BAD", "NYY"), ("ATL", "NYM")]),
                        ("b.json", [("BOS", "TBR")])):
        p = tmp_path / name
        games = [dict(_spec(a, h), start_et="19:05")
                 for a, h in teams]
        p.write_text(json.dumps({"games": games}), encoding="utf-8")
        slates.append(str(p))
    saved = []
    monkeypatch.setattr(
        PR, "save_excel_slate",
        lambda specs, out, ledger=True:
            saved.append([r["degraded"] for r in out])
            or str(tmp_path / "wb.xlsx"))
    monkeypatch.setattr(PR, "_serve_health_warning", lambda: None)
    monkeypatch.setattr(PR, "Predictor", lambda progress=None: P)
    monkeypatch.setattr(sys, "argv",
                        ["predict.py", "--serve", "--json"] + slates)
    with pytest.raises(SystemExit) as ei:
        PR.main()
    assert ei.value.code == 3
    assert len(saved) == 2
    # slate a's workbook carries its own failure...
    assert any("FAILED" in d for degs in saved[0] for d in degs)
    # ...and slate b's workbook is pristine
    assert all(degs == [] for degs in saved[1])
