"""Serve path: slate specs -> simulated games -> workbook.

Predictor loads the trained artifacts and the feature warehouse once,
then predict_slate() prices any list of game specs (the GUI's or the
backtester's):

  1. resolve the spec into player rows (9 batters a side from the posted
     lineup, filled from the club's last posted order when short;
     starters; the bullpen from roster + recent-usage availability)
  2. assemble as-of features for every (pitcher, batter, times-through-
     order) matchup through features.assemble_features — the SAME code
     path training used — and evaluate the calibrated A1/A2 models
  3. build the starter hazard grids, steal matrices, pattern bank and
     latent sigmas into a sim.GamePrep
  4. run the vectorized sim and count every market off the tensor

save_excel_slate() writes the workbook the Tools expect (sheets: Batter
Props, Pitching Props, Games, Bets) to Predictions/. Per-sim tensor
persistence and the SGP pricer (Model/sgp.py) were removed 2026-07-19
by user decision (no parlay betting) — both recoverable from git
history if SGP pricing is ever wanted.
"""

import datetime as dt
import json
import os
from pathlib import Path

import joblib
import numpy as np
import pandas as pd

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
import features as F     # noqa: E402
import sim               # noqa: E402
import odds as O         # noqa: E402

# canonical context-feature names for the residual heads; heads.py
# aliases this list, so train-time and serve-time can never diverge
HEAD_CTX = ["elo", "rest", "b2b", "travel_km", "tz_shift",
            "day_after_night", "w10", "rf10", "ra10", "pyth", "gp",
            "pen_era", "pen_np3", "st_outs", "bsr_luck"]

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "Data"
ART = ROOT / "Model" / "artifacts"
PRED_DIR = ROOT / "Predictions"

N_SIMS = 20000
MAX_PEN = 8
TIRED_NP = 20            # a pen arm that threw >= this many pitches
                         # YESTERDAY drops behind every fresh arm in
                         # both entry orders (threw-both-of-last-two-
                         # days / 25+ NP arms are excluded upstream)
PEN_FAT_MID = 10         # B3 graded demotion: 10-19 NP yesterday is a
                         # middle tier between fresh and TIRED_NP
PEN_EXIT_M = 12          # B4 per-pitcher exit tables: league pseudo-
                         # stints blended into the pitcher's trailing-
                         # 365d stint hazard (holdout-tuned,
                         # Logs/pen_exit_study_2026-07-20.log)
# A/B gates for the 2026-07-20 pen wave: set to "0" to reproduce the
# exact pre-wave behavior (paired-replay A sides). PEN_WAVE3 covers
# B3 (fatigue offsets + graded tiers) + B4 (per-pitcher exit tables);
# PEN_CHOICE covers B6 (pmf-sampled reliever entry rank).
PEN_WAVE3 = os.environ.get("PEN_WAVE3", "1") != "0"
PEN_CHOICE = os.environ.get("PEN_CHOICE", "1") != "0"

K_LINES = [3.5, 4.5, 5.5, 6.5, 7.5, 8.5]
OUT_LINES = [14.5, 15.5, 16.5, 17.5, 18.5]
PHA_LINES = [3.5, 4.5, 5.5, 6.5]
PBB_LINES = [0.5, 1.5, 2.5]
PER_LINES = [1.5, 2.5, 3.5, 4.5]
TOTAL_LINES = [6.5, 7.5, 8.5, 9.5, 10.5]
TEAM_TOTAL_LINES = [2.5, 3.5, 4.5, 5.5]

# workbook look (matches the established output format; 8-digit ARGB so
# the stored colors byte-match the format the grader expects)
NAVY = "FF041E42"
BETS_HEADER = "FF1E7B34"
BETS_ROW_TINT = "FFE7F3E2"      # the Bets rows (grading paints winners)
ROOKIE_RED = "FFC00000"
ROOKIE_G = 50                   # career games under this -> red + note
MIN_EV = 0.05                   # Bets sheet inclusion floor

BAT_BINARIES = [
    ("HR", lambda t, s: t[..., s["HR"]] >= 1),
    ("Hit", lambda t, s: t[..., s["H"]] >= 1),
    ("2+ Hits", lambda t, s: t[..., s["H"]] >= 2),
    ("Single", lambda t, s: t[..., s["B1"]] >= 1),
    ("Double", lambda t, s: t[..., s["B2"]] >= 1),
    ("Triple", lambda t, s: t[..., s["B3"]] >= 1),
    ("2+ TB", lambda t, s: _tb(t, s) >= 2),
    ("3+ TB", lambda t, s: _tb(t, s) >= 3),
    ("4+ TB", lambda t, s: _tb(t, s) >= 4),
    ("Run", lambda t, s: t[..., s["R"]] >= 1),
    ("2+ Runs", lambda t, s: t[..., s["R"]] >= 2),
    ("RBI", lambda t, s: t[..., s["RBI"]] >= 1),
    ("2+ RBI", lambda t, s: t[..., s["RBI"]] >= 2),
    ("H+R+RBI 2+", lambda t, s: _hrr(t, s) >= 2),
    ("H+R+RBI 3+", lambda t, s: _hrr(t, s) >= 3),
    ("H+R+RBI 4+", lambda t, s: _hrr(t, s) >= 4),
    ("BB", lambda t, s: t[..., s["BB"]] >= 1),
    ("SB", lambda t, s: t[..., s["SB"]] >= 1),
    ("K", lambda t, s: t[..., s["K"]] >= 1),
    ("2+ K", lambda t, s: t[..., s["K"]] >= 2),
    ("3+ K", lambda t, s: t[..., s["K"]] >= 3),
]


def _tb(t, s):
    return (t[..., s["B1"]] + 2 * t[..., s["B2"]] + 3 * t[..., s["B3"]]
            + 4 * t[..., s["HR"]])


def _hrr(t, s):
    return t[..., s["H"]] + t[..., s["R"]] + t[..., s["RBI"]]


PB_K = 64.0     # pseudo-sims of parametric-tail shrinkage


def _tail_prob(counts, thr):
    """P(count > thr) with a parametric shrink for sparse tails:
    (empirical hits + PB_K * parametric) / (n + PB_K), the parametric
    term a moment-fit negative binomial (Poisson when var <= mean).
    Data-rich probabilities stay essentially empirical; beyond-support
    lines get smooth nonzero tails instead of hard 0s."""
    from scipy import stats as _st
    counts = np.asarray(counts)
    n = counts.size
    hits = float((counts > thr).sum())
    m = float(counts.mean())
    v = float(counts.var())
    k = int(np.floor(thr))
    if m <= 1e-9:
        pp = 0.0
    elif v > m * 1.0001:
        r = m * m / (v - m)
        pp = float(_st.nbinom.sf(k, r, r / (r + m)))
    else:
        pp = float(_st.poisson.sf(k, m))
    return float((hits + PB_K * pp) / (n + PB_K))


class _Stores:
    def __init__(self, raw):
        self.raw = raw


def _load_raw():
    """The GUI-facing raw frames (contract: games/rosters/gb/gp/parks/
    umps)."""
    games = pd.read_csv(DATA / "mlb_games.csv", encoding="utf-8-sig",
                        low_memory=False)
    games["Date"] = pd.to_datetime(games["Date"])
    ros = pd.read_csv(DATA / "mlb_rosters.csv", encoding="utf-8-sig")
    gb = pd.read_csv(DATA / "mlb_game_batting.csv", encoding="utf-8-sig",
                     low_memory=False,
                     usecols=["GamePk", "Season", "Date", "PlayerId",
                              "Name", "Team", "BattingOrder", "Position",
                              "PA"])
    gb["Date"] = pd.to_datetime(gb["Date"])
    gp = pd.read_csv(DATA / "mlb_game_pitching.csv", encoding="utf-8-sig",
                     low_memory=False,
                     usecols=["GamePk", "Season", "Date", "PlayerId",
                              "Name", "Team", "GS", "GF", "IP", "BF",
                              "NP", "SV", "HLD"])
    gp["Date"] = pd.to_datetime(gp["Date"])
    parks = pd.read_csv(DATA / "mlb_ballparks.csv", encoding="utf-8-sig")
    umps = pd.read_csv(DATA / "mlb_umpires.csv", encoding="utf-8-sig")
    return {"games": games, "rosters": ros, "gb": gb, "gp": gp,
            "parks": parks, "umps": umps}


def spec_postseason(spec):
    """Postseason flag for the SB models: explicit game_type when the
    spec carries it (replays stamp it from mlb_games), else the October
    heuristic for live slates (wild-card games in late September before
    the schedule scrape catches up are the only miss)."""
    gt = str(spec.get("game_type") or "")
    if gt:
        return float(gt in ("F", "D", "L", "W"))
    try:
        return float(pd.Timestamp(spec["date"]).month >= 10)
    except (KeyError, ValueError):
        return 0.0


class Predictor:
    def __init__(self, progress=None):
        tick = progress or (lambda m: None)
        tick("loading artifacts...")
        # artifacts trained by `python Model/train.py` pickled the scaler
        # under __main__; resolve it for those runs
        import __main__ as _m
        if not hasattr(_m, "VectorScaler"):
            _m.VectorScaler = F.VectorScaler
        self.a1 = joblib.load(ART / "a1_model.joblib")
        self.a2 = joblib.load(ART / "a2_model.joblib")
        self.hz = joblib.load(ART / "hazard_model.joblib")
        self.sb = joblib.load(ART / "sb_models.joblib")
        self.latent = json.loads((ART / "latent.json").read_text())
        tick("loading feature stores...")
        self.fstores = F.load_stores()
        self.patterns = sim.load_patterns(F.STORES)
        self.preevents = json.loads(
            (F.STORES / "preevents.json").read_text())
        tick("loading data frames...")
        self.stores = _Stores(_load_raw())
        self._names = {}
        for df, idc, nc in ((self.stores.raw["gb"], "PlayerId", "Name"),
                            (self.stores.raw["gp"], "PlayerId", "Name"),
                            (self.stores.raw["rosters"], "PlayerId",
                             "Name")):
            self._names.update(dict(zip(df[idc], df[nc])))
        gb_ = self.stores.raw["gb"]
        self._career_g = gb_[pd.to_numeric(gb_.PA, errors="coerce") > 0] \
            .groupby("PlayerId").size().to_dict()
        gp_ = self.stores.raw["gp"]
        self._career_gp = gp_.groupby("PlayerId").size().to_dict()
        hand = pd.read_csv(DATA / "mlb_handedness.csv",
                           encoding="utf-8-sig")
        self._bats = dict(zip(hand.PlayerId, hand.Bats))
        self._throws = dict(zip(hand.PlayerId, hand.Throws))
        ros = self.stores.raw["rosters"]
        self._bats.update({p: b for p, b in zip(ros.PlayerId, ros.B)
                           if p not in self._bats})
        self._throws.update({p: t for p, t in zip(ros.PlayerId, ros.T)
                             if p not in self._throws})
        # club abbrev -> full name -> roster rows
        bs = pd.read_csv(DATA / "mlb_batting_stats.csv",
                         encoding="utf-8-sig",
                         usecols=["Year", "Team", "TeamName"])
        pairs = bs[bs.Year == bs.Year.max()].drop_duplicates()
        self._full_of = dict(zip(pairs.Team, pairs.TeamName))
        ros = self.stores.raw["rosters"]
        self._catchers = set(
            ros.loc[ros.Position == "Catcher", "PlayerId"].astype(int))
        self._roster_pos = dict(zip(ros.PlayerId.astype(int),
                                    ros.Position))
        # batter participation hazard table -> dense lookup array
        self.part_haz = None
        part_path = F.STORES / "participation.parquet"
        if part_path.exists():
            pt = pd.read_parquet(part_path)
            km = pt.groupby("k").agg(ev=("ev", "sum"), n=("n", "sum"))
            km = (km.ev / km.n.clip(lower=1)).to_dict()
            dense = np.zeros((6, 4, 3, 2, 2, 2))
            for kk in range(1, 7):
                dense[kk - 1, ...] = km.get(kk, 0.0)
            for r in pt.itertuples(index=False):
                dense[int(r.k) - 1, int(r.inn_b), int(r.margin_b),
                      int(r.lead), int(r.same), int(r.isc)] = r.rate
            self.part_haz = dense
        # cross-line-coherent output calibrators (one shared monotone
        # map per market FAMILY; fit by evaluate.py --fit-calibrators)
        cal_path = ART / "output_calibrators.joblib"
        self.calib = joblib.load(cal_path) if cal_path.exists() else {}
        # residual heads (heads.py --train): per-family shallow-GBM
        # logit corrections applied on top of the Platt output. Only
        # heads that kept trees load; to serve Platt-only, move
        # residual_heads.joblib aside.
        self.heads = {}
        hp = ART / "residual_heads.joblib"
        if hp.exists():
            import lightgbm as lgb
            for fam, h in joblib.load(hp).items():
                if h.get("best_iter", 0) > 0:
                    self.heads[fam] = dict(
                        bst=lgb.Booster(model_str=h["booster_str"]),
                        features=h["features"],
                        best_iter=int(h["best_iter"]))
        self._defense = {}
        dp = F.STORES / "defense_team.parquet"
        if dp.exists():
            dt_ = pd.read_parquet(dp)
            self._defense = {
                (int(y), str(t)): (o, fr) for y, t, o, fr in zip(
                    dt_.Year, dt_.Team, dt_.def_oaa, dt_.FrameRV_pt)}
        # sim identity-adjuster stores (features.py --only simadj);
        # every one degrades to neutral engine behavior when absent
        def _pq(name):
            p = F.STORES / name
            return pd.read_parquet(p) if p.exists() else None

        self._runner_z, self._stretch_zmap = {}, {}
        self._catcher_wpmap, self._arm_of_p, self._arm_of_t = {}, {}, {}
        rz_ = _pq("runner_z.parquet")
        if rz_ is not None:
            self._runner_z = {
                (int(s), int(p)): (float(r), float(d)) for s, p, r, d in
                zip(rz_.Season, rz_.PlayerId, rz_.run_z, rz_.dp_z)}
        am_ = _pq("arm_of.parquet")
        if am_ is not None:
            self._arm_of_p = {(int(s), int(p)): float(z) for s, p, z in
                              zip(am_.Season, am_.PlayerId, am_.arm_z)}
        at_ = _pq("arm_of_team.parquet")
        if at_ is not None:
            self._arm_of_t = {(int(s), str(t)): float(z) for s, t, z in
                              zip(at_.Season, at_.Team, at_.team_arm_z)}
        sz_ = _pq("stretch_z.parquet")
        if sz_ is not None:
            self._stretch_zmap = {
                (int(s), int(p)): float(z) for s, p, z in
                zip(sz_.Season, sz_.PlayerId, sz_.stretch_z)}
        cw_ = _pq("catcher_wp.parquet")
        if cw_ is not None:
            self._catcher_wpmap = {
                (int(s), int(p)): float(r) for s, p, r in
                zip(cw_.Season, cw_.PlayerId, cw_.wp_rate)}
        tj = F.STORES / "advance_tilt.json"
        self._tilt = json.loads(tj.read_text()) if tj.exists() else {}
        sj = F.STORES / "stretch.json"
        self._stretch = json.loads(sj.read_text()) if sj.exists() \
            else None
        self._outfielders = set(ros.loc[ros.Position.isin(
            ["Left Field", "Center Field", "Right Field"]),
            "PlayerId"].astype(int))
        tick("computing bullpen availability...")
        self._relief_exit = self._reliever_exit_table()
        # per-pitcher relief stint outs (B4: pitcher-specific exit
        # tables) and fatigue coefficients (B3: only |z|>=2 classes
        # apply; the store records the full fit)
        gp_rel = gp_[pd.to_numeric(gp_.GS, errors="coerce") == 0]
        ipn = pd.to_numeric(gp_rel.IP, errors="coerce").fillna(0)
        outs_ = (ipn.astype(int) * 3
                 + np.round((ipn % 1) * 10)).astype(int)
        self._rel_outs = {
            int(p): (g.Date.to_numpy(), g.o.to_numpy())
            for p, g in pd.DataFrame(dict(
                PlayerId=gp_rel.PlayerId, Date=gp_rel.Date,
                o=outs_)).groupby("PlayerId")}
        self._pexit_cache = {}
        self._pen_fat = {}
        pf = F.STORES / "pen_fatigue.json"
        if PEN_WAVE3 and pf.exists():
            d_ = json.loads(pf.read_text())
            ci = {"K": 0, "BB": 1, "HR": 6}
            self._pen_fat = {
                ci[c]: float(b) for c, b in d_["beta"].items()
                if c in ci and abs(d_["z"].get(c, 0.0)) >= 2.0}
        # B6: cumulative rank pmfs [hi, lo] for the sim's entry pick
        self._pen_rank_cum = None
        pc = F.STORES / "pen_choice.json"
        if PEN_CHOICE and pc.exists():
            d_ = json.loads(pc.read_text())
            if d_.get("n_hi", 0) >= 500 and d_.get("n_lo", 0) >= 500:
                self._pen_rank_cum = np.cumsum(
                    np.asarray([d_["hi"], d_["lo"]], dtype=np.float64),
                    axis=1)
        tick("ready")

    # ------------------------------------------------------ helpers

    def _name(self, pid):
        return self._names.get(pid, str(pid))

    def _stand(self, bat_pid, p_throws):
        b = str(self._bats.get(bat_pid, "R"))
        if b == "S":
            return "L" if p_throws == "R" else "R"
        return b if b in ("L", "R") else "R"

    def _reliever_exit_table(self):
        gp = self.stores.raw["gp"]
        rel = gp[(pd.to_numeric(gp.GS, errors="coerce") == 0)].copy()
        ip = pd.to_numeric(rel.IP, errors="coerce").fillna(0)
        outs = (ip.astype(int) * 3 + round((ip % 1) * 10)).astype(int)
        tab = np.ones(11)
        for k in range(1, 11):
            at_least = (outs >= k).sum()
            more = (outs > k).sum()
            tab[k] = 1.0 - (more / max(at_least, 1))
        tab[0] = 0.0
        # exit decisions happen at inning breaks; the per-outs exit prob
        # is evaluated there with the stint's completed outs
        return tab

    def _pitcher_exit_table(self, pid, date):
        """[11] per-outs exit probs for one reliever: trailing-365d
        stint hazard counts blended with PEN_EXIT_M league pseudo-
        stints (B4 — closers exit after 3 outs, long men carry
        multi-inning stints)."""
        key = (int(pid), pd.Timestamp(date))
        hit = self._pexit_cache.get(key)
        if hit is not None:
            return hit
        tab = self._relief_exit
        arr = self._rel_outs.get(int(pid))
        if arr is not None:
            d = np.datetime64(pd.Timestamp(date))
            dates, outs = arr
            o = outs[(dates >= d - np.timedelta64(365, "D"))
                     & (dates < d)]
            if len(o):
                k = np.arange(11)
                at_least = (o[:, None] >= k).sum(0)
                exits = at_least - (o[:, None] > k).sum(0)
                tab = ((exits + PEN_EXIT_M * tab)
                       / (at_least + PEN_EXIT_M))
                tab[0] = 0.0
        self._pexit_cache[key] = tab
        return tab

    def _apply_pen_fatigue(self, avec, pen_fat_f):
        """B3 performance half: multiply fatigued pen arms' class
        probabilities by exp(beta * f) for the store's significant
        classes (BB as of the 2026-07-20 fit) and renormalize."""
        if not self._pen_fat or not pen_fat_f:
            return
        for r, f in pen_fat_f.items():
            if f <= 0:
                continue
            v = avec[r]
            for ci, b in self._pen_fat.items():
                v[..., ci] *= np.exp(b * f)
            v /= v.sum(axis=-1, keepdims=True)

    def _il_unavailable(self, date):
        """PlayerId -> IL PlaceDate for stints active on `date` (as-of
        safe: placements are announced pregame). Open stints (no
        ActDate — season-enders never get one) stay active; a game-log
        appearance after PlaceDate overrides in _pen_for."""
        il = getattr(self, "_il_cache", None)
        if il is None:
            try:
                il = pd.read_csv(DATA / "mlb_il.csv",
                                 encoding="utf-8-sig",
                                 usecols=["PlayerId", "PlaceDate",
                                          "ActDate"])
                il["PlaceDate"] = pd.to_datetime(il["PlaceDate"])
                il["ActDate"] = pd.to_datetime(il["ActDate"])
            except FileNotFoundError:
                il = pd.DataFrame(columns=["PlayerId", "PlaceDate",
                                           "ActDate"])
            self._il_cache = il
            self._il_on_cache = {}
        d = pd.Timestamp(date)
        if d not in self._il_on_cache:
            on = il[(il.PlaceDate <= d)
                    & (il.ActDate.isna() | (il.ActDate > d))]
            self._il_on_cache[d] = dict(
                zip(on.PlayerId.astype(int), on.PlaceDate))
        return self._il_on_cache[d]

    def _pit_team_last(self, date):
        """PlayerId -> (team, last appearance Date) across ALL teams in
        the 30 days before `date` — catches mid-window trades that the
        per-team usage logs can't see."""
        cache = getattr(self, "_ptl_cache", None)
        if cache is None:
            cache = {}
            self._ptl_cache = cache
        d = pd.Timestamp(date)
        if d not in cache:
            gp = self.stores.raw["gp"]
            win = gp[(gp.Date >= d - pd.Timedelta(days=30))
                     & (gp.Date < d)]
            last = win.sort_values("Date").groupby("PlayerId").tail(1)
            cache[d] = {int(p): (t, dt) for p, t, dt in
                        zip(last.PlayerId, last.Team, last.Date)}
        return cache[d]

    def _pen_for(self, team_abbrev, date):
        """Available relievers with usage weights: roster bullpen minus
        arms that threw on both of the last two days (or 25+ pitches
        yesterday), minus arms on the IL or last seen pitching for
        another club (historical replays reconstruct availability from
        IL spans + game logs; the roster snapshot only exists for
        today)."""
        cache = getattr(self, "_gp_team_cache", None)
        if cache is None:
            gp_all = self.stores.raw["gp"]
            cache = {t: sub.sort_values("Date")
                     for t, sub in gp_all.groupby("Team")}
            self._gp_team_cache = cache
        gp = cache.get(team_abbrev,
                       self.stores.raw["gp"].iloc[0:0])
        ros = self.stores.raw["rosters"]
        full = self._full_of.get(team_abbrev)
        pen_ids = ros.loc[(ros.Team == full)
                          & (ros.Position == "Bullpen"), "PlayerId"]
        pen_ids = [int(p) for p in pen_ids]
        d = pd.Timestamp(date)
        recent = gp[(gp.Date >= d - pd.Timedelta(days=30))
                    & (gp.Date < d)
                    & (pd.to_numeric(gp.GS, errors="coerce") == 0)]
        y1 = recent[recent.Date == d - pd.Timedelta(days=1)]
        y2 = recent[recent.Date == d - pd.Timedelta(days=2)]
        np1 = dict(zip(y1.PlayerId, pd.to_numeric(y1.NP,
                                                  errors="coerce")))
        used_both = set(y1.PlayerId) & set(y2.PlayerId)
        out = []
        counts = recent.groupby("PlayerId").size()
        il_on = self._il_unavailable(d)
        team_last = self._pit_team_last(d)
        last_app = recent.groupby("PlayerId").Date.max()

        def _available(pid):
            place = il_on.get(pid)
            if place is not None and not (
                    pid in last_app.index and last_app[pid] >= place):
                return False        # on IL, nothing pitched since
            tl = team_last.get(pid)
            if tl is not None and tl[0] != team_abbrev:
                return False        # last seen with another club
            return True

        for pid in pen_ids:
            if pid in used_both or np1.get(pid, 0) >= 25:
                continue
            if not _available(pid):
                continue
            out.append((pid, 1.0 + counts.get(pid, 0)))
        # game-log fallback when the depth chart is thin
        if len(out) < 5:
            for pid, n in counts.sort_values(ascending=False).items():
                if pid not in {p for p, _ in out} and len(out) < MAX_PEN \
                        and _available(int(pid)):
                    out.append((int(pid), 1.0 + n))
        out.sort(key=lambda t: -t[1])
        return out[:MAX_PEN]

    def _class_vecs(self, X):
        """A1/A2 probabilities for assembled matchup rows. Returns
        (p8 [n,8], a2arr): a2arr is [n,4] P(bb | in-play) under the
        flat artifact or [n,8,4] P(bb | outcome class) under the
        hierarchical contact tree."""
        X2 = X.reindex(columns=self.a2["features"]).astype(np.float32)
        p2 = self.a2["scaler"].transform(
            self.a2["model"].predict_proba(X2))
        X1 = X.reindex(columns=self.a1["features"]).astype(np.float32)
        if self.a1.get("kind") != "tree":
            p8 = self.a1["scaler"].transform(
                self.a1["model"].predict_proba(X1))
            return p8, p2
        t1, t3 = self.a1["t1"], self.a1["t3"]
        p1 = t1["scaler"].transform(t1["model"].predict_proba(X1))
        Xn = X1.to_numpy()

        def t3mat(bi):
            d = np.zeros((len(Xn), 3), dtype=np.float32)
            if bi > 0:
                d[:, bi - 1] = 1.0
            return np.column_stack([Xn, d])

        p3 = np.stack(
            [t3["scaler"].transform(t3["model"].predict_proba(t3mat(bi)))
             for bi in range(4)], axis=1)
        return F.tree_compose(p1, p2, p3)

    _POS_SLOT = {"First Base": 3, "Second Base": 4, "Third Base": 5,
                 "Shortstop": 6, "Left Field": 7, "Center Field": 8,
                 "Right Field": 9}

    def _fielder_map(self, pids):
        """Lineup pids -> {savant fielder slot 3-9: pid} by roster
        primary position (first claimant wins; unfilled slots NaN)."""
        fm = {i: np.nan for i in range(3, 10)}
        for pid in pids:
            if pid is None or pid < 0:
                continue
            sl = self._POS_SLOT.get(str(self._roster_pos.get(int(pid))))
            if sl and not (fm[sl] == fm[sl]):
                fm[sl] = int(pid)
        return fm

    def _sim_adjusters(self, players, season, away, home, date,
                       pen_rows_away, pen_rows_home):
        """Identity-adjuster prep arrays for one game: runner/GIDP z per
        player row, starter stretch z, OF-arm advancement effect and
        catcher WP+PB rate per FIELDING side, and the leverage-aware pen
        entry orders (high-leverage: save/hold arms first; low-leverage:
        length first)."""
        n = len(players)
        run_z = np.zeros(n, dtype=np.float32)
        dp_z = np.zeros(n, dtype=np.float32)
        stretch_z = np.zeros(n, dtype=np.float32)
        for i, pid in enumerate(players):
            if pid is None or pid < 0:
                continue
            if i < 18:
                rz = self._runner_z.get((season, int(pid)))
                if rz:
                    run_z[i], dp_z[i] = rz
            else:
                stretch_z[i] = self._stretch_zmap.get(
                    (season, int(pid)), 0.0)

        th = self._tilt
        ratio = (th.get("theta_arm", 0.0) / th["theta_adv"]
                 if th.get("theta_adv") else 0.0)
        arm_eff = np.zeros(2, dtype=np.float32)
        pre_wp = np.full(
            2, float(self.preevents["wp_pb_per_pa_runners_on"]),
            dtype=np.float32)
        for side, team, lu in ((0, away, players[0:9]),
                               (1, home, players[9:18])):
            zs = [self._arm_of_p[(season, int(p))] for p in lu
                  if p >= 0 and int(p) in self._outfielders
                  and (season, int(p)) in self._arm_of_p]
            tz = (float(np.mean(zs)) if zs
                  else self._arm_of_t.get((season, team), 0.0))
            arm_eff[side] = ratio * tz
            cid = next((int(p) for p in lu
                        if p >= 0 and int(p) in self._catchers), None)
            if cid is not None:
                pre_wp[side] = self._catcher_wpmap.get(
                    (season, cid), pre_wp[side])

        pen_hi = np.full((2, MAX_PEN), -1, dtype=np.int16)
        pen_lo = np.full((2, MAX_PEN), -1, dtype=np.int16)
        rex = np.tile(self._relief_exit.astype(np.float32),
                      (len(players), 1))
        pen_fat_f = {}
        gp_cache = getattr(self, "_gp_team_cache", {})
        d = pd.Timestamp(date)
        for side, team, prows in ((0, away, pen_rows_away),
                                  (1, home, pen_rows_home)):
            stats = {}
            sub = gp_cache.get(team)
            if sub is not None and prows:
                rec = sub[(sub.Date >= d - pd.Timedelta(days=365))
                          & (sub.Date < d)
                          & (pd.to_numeric(sub.GS,
                                           errors="coerce") == 0)]
                if len(rec):
                    ipn = pd.to_numeric(rec.IP, errors="coerce").fillna(0)
                    fr = pd.DataFrame(dict(
                        pid=rec.PlayerId,
                        sv=pd.to_numeric(rec.SV,
                                         errors="coerce").fillna(0),
                        hld=pd.to_numeric(rec.HLD,
                                          errors="coerce").fillna(0),
                        outs=(ipn.astype(int) * 3
                              + np.round((ipn % 1) * 10))))
                    g = fr.groupby("pid").agg(sv=("sv", "sum"),
                                              hld=("hld", "sum"),
                                              outs=("outs", "mean"))
                    stats = {int(p): (float(r.sv), float(r.hld),
                                      float(r.outs))
                             for p, r in g.iterrows()}
            # B3 graded demotion: recent pitch counts split arms into
            # fresh / mid (PEN_FAT_MID+) / heavy (TIRED_NP+ or went
            # back-to-back) tiers; each tier sorts behind the previous
            # one in both entry orders (managers hold worked arms back
            # short of the true unavailability _pen_for handles)
            np1, np2 = {}, {}
            if sub is not None and prows:
                gs0 = pd.to_numeric(sub.GS, errors="coerce") == 0
                for lag, m_ in ((1, np1), (2, np2)):
                    yd = sub[(sub.Date == d - pd.Timedelta(days=lag))
                             & gs0]
                    if len(yd):
                        npv = pd.to_numeric(yd.NP,
                                            errors="coerce").fillna(0)
                        for p, v in zip(yd.PlayerId, npv):
                            m_[int(p)] = m_.get(int(p), 0.0) + float(v)

            def _tier(pid):
                n1, n2 = np1.get(pid, 0.0), np2.get(pid, 0.0)
                if not PEN_WAVE3:               # legacy binary demotion
                    return 1 if n1 >= TIRED_NP else 0
                if n1 >= TIRED_NP or (n1 > 0 and n2 > 0):
                    return 2
                return 1 if n1 >= PEN_FAT_MID else 0

            rows = [(r, *stats.get(int(players[r]), (0.0, 0.0, 3.0)),
                     _tier(int(players[r])))
                    for r in prows]
            hi = sorted(rows,
                        key=lambda t: (t[4], -(2.0 * t[1] + t[2])))
            lo = sorted(rows,
                        key=lambda t: (t[4],
                                       -(t[3] - 0.5 * (t[1] + t[2]))))
            pen_hi[side, :len(hi)] = [r for r, *_ in hi]
            pen_lo[side, :len(lo)] = [r for r, *_ in lo]
            if PEN_WAVE3:
                for r in prows:
                    pid = int(players[r])
                    n1, n2 = np1.get(pid, 0.0), np2.get(pid, 0.0)
                    pen_fat_f[r] = (min(n1, 40.0) / 20.0
                                    + 0.5 * min(n2, 40.0) / 20.0)
                    if pid > 0:
                        rex[r] = self._pitcher_exit_table(pid, d)

        return dict(run_z=run_z, dp_z=dp_z, stretch_z=stretch_z,
                    arm_eff=arm_eff, pre_wp=pre_wp, pen_hi=pen_hi,
                    pen_lo=pen_lo, stretch=self._stretch,
                    relief_exit=(rex if PEN_WAVE3
                                 else self._relief_exit),
                    pen_fat_f=pen_fat_f)

    def _last_lineup(self, team):
        gb = self.stores.raw["gb"]
        rows = gb[(gb.Team == team) & gb.BattingOrder.notna()].copy()
        rows["bo"] = pd.to_numeric(rows.BattingOrder, errors="coerce")
        rows = rows[rows.bo % 100 == 0]
        if rows.empty:
            return []
        last = rows[rows.Date == rows.Date.max()].sort_values("bo")
        return [int(p) for p in last.PlayerId.head(9)]

    # -------------------------------------------------- game assembly

    def prepare_game(self, spec, n_sims=N_SIMS):
        date = pd.Timestamp(spec["date"])
        season = int(spec.get("season") or date.year)
        away, home = spec["away_team"], spec["home_team"]

        def side_lineup(side, team):
            posted = sorted(spec.get(f"{side}_lineup") or [],
                            key=lambda x: x[1])
            pids = [int(p) for p, _ in posted][:9]
            if len(pids) < 9:
                for p in self._last_lineup(team):
                    if p not in pids:
                        pids.append(p)
                    if len(pids) == 9:
                        break
            while len(pids) < 9:
                pids.append(-1)          # league-average phantom
            return pids

        bat_away = side_lineup("away", away)
        bat_home = side_lineup("home", home)
        sp_away = int(spec.get("away_starter") or -2)
        sp_home = int(spec.get("home_starter") or -3)
        pen_away = self._pen_for(away, date)
        pen_home = self._pen_for(home, date)

        # rows: 0-17 lineups, 18/19 starters, 20-35 pens, 36/37 = the
        # generic BENCH bats substitutions bring in (league-average
        # phantoms; the sim maps them to batter-axis slots 18/19)
        players = (bat_away + bat_home + [sp_away, sp_home]
                   + [p for p, _ in pen_away]
                   + [-1] * (MAX_PEN - len(pen_away))
                   + [p for p, _ in pen_home]
                   + [-1] * (MAX_PEN - len(pen_home))
                   + [-1, -1])
        n_players = len(players)         # 18 + 2 + 16 + 2 bench
        bench_rows = (n_players - 2, n_players - 1)
        pit_rows = [18, 19] + list(range(20, 20 + MAX_PEN)) \
            + list(range(20 + MAX_PEN, 20 + 2 * MAX_PEN))
        pen_rows_away = list(range(20, 20 + len(pen_away)))
        pen_rows_home = list(range(20 + MAX_PEN,
                                   20 + MAX_PEN + len(pen_home)))
        bat_rows_all = list(range(18)) + list(bench_rows)

        # pre-game team context (Elo) + the fielding team's actual
        # fielders by roster position — A1 features; ctx is injected by
        # predict_slate (spec["_ctx"]); absent -> NaN (XGB-native)
        ctx = spec.get("_ctx") or {}
        elo_a = ctx.get("away_elo", np.nan)
        elo_h = ctx.get("home_elo", np.nan)
        cx_b = {}   # schedule-spot A1 features, keyed by bat_is_home
        for hb in (0, 1):
            own, opp = ("home", "away") if hb else ("away", "home")
            cx_b[hb] = {
                f"{pre}_{c}": ctx.get(f"{sd}_{c}", np.nan)
                for c in ("travel_km", "tz_shift", "day_after_night")
                for pre, sd in (("b", own), ("p", opp))}
        fm_side = {away: self._fielder_map(bat_away),
                   home: self._fielder_map(bat_home)}

        # ---- matchup feature rows for A1/A2 (18 lineup + 2 bench bats)
        rows = []
        for prow in pit_rows:
            ppid = players[prow]
            pthrows = str(self._throws.get(ppid, "R"))
            pthrows = pthrows if pthrows in ("L", "R") else "R"
            p_team = away if (prow == 18 or prow in pen_rows_away) \
                else home
            for brow in bat_rows_all:
                # rows exist for every (pitcher, batter) pair so the
                # avec indexing stays rectangular; the sim only ever
                # reads cross-team pairs
                bpid = players[brow]
                bat_is_home = (brow >= 9) if brow < 18 else \
                    (brow == bench_rows[1])
                stand = self._stand(bpid, pthrows)
                # on-deck = next lineup slot, same side (bench -> none)
                od_pid = (players[(brow + 1) % 9] if brow < 9 else
                          players[9 + (brow - 8) % 9] if brow < 18
                          else -1)
                for tto in (1, 2, 3):
                    rows.append(dict(
                        Date=date, Season=season, BatterId=bpid,
                        PitcherId=ppid, stand=stand, p_throws=pthrows,
                        OnDeckId=od_pid,
                        tto=tto, home_bat=int(bat_is_home),
                        fld_team=p_team, Venue=spec.get("venue") or "",
                        DayNight=spec.get("day_night") or "day",
                        Temp=spec.get("temp"),
                        WindSpeed=spec.get("wind_speed"),
                        WindDir=spec.get("wind_dir") or "",
                        Condition=spec.get("condition") or "",
                        Humidity=spec.get("humidity"),
                        Pressure=spec.get("pressure"),
                        HpUmpId=spec.get("hp_ump_id"),
                        rest_p=self._rest(ppid, date),
                        b_elo=(elo_h if bat_is_home else elo_a),
                        p_elo=(elo_a if bat_is_home else elo_h),
                        **cx_b[int(bat_is_home)],
                        **{f"fielder_{i}": fm_side[p_team][i]
                           for i in range(3, 10)},
                    ))
        rdf = pd.DataFrame(rows)
        X, _ = F.assemble_features(rdf, self.fstores)
        p1, p2 = self._class_vecs(X)
        avec = np.full((n_players, 20, 3, 8), np.nan)
        a2vec = np.full((n_players, 20, 3) + p2.shape[1:], np.nan)
        i = 0
        for prow in pit_rows:
            for bi, brow in enumerate(bat_rows_all):
                for t in range(3):
                    avec[prow, bi, t] = p1[i]
                    a2vec[prow, bi, t] = p2[i]
                    i += 1

        # ---- hazard grids for the two starters
        haz = np.zeros((2, 41, 11))
        post_g = float(spec_postseason(spec))
        for si, (prow, ppid, team) in enumerate(
                ((18, sp_away, away), (19, sp_home, home))):
            haz[si] = self._hazard_grid(ppid, date, season, team,
                                        post=post_g)

        # ---- steal matrices (runner PLAYER ROW vs every pitcher row)
        sb_att, sb_suc = self._sb_matrices(players, pit_rows,
                                           bat_rows_all, date, season,
                                           away, home,
                                           post=spec_postseason(spec))
        sb_state = self._sb_state(season)

        pen_order = np.zeros((1, 2, MAX_PEN), dtype=np.int16) - 1
        pen_order[0, 0, :len(pen_rows_away)] = pen_rows_away
        pen_order[0, 1, :len(pen_rows_home)] = pen_rows_home
        pen_order = np.broadcast_to(pen_order, (n_sims, 2, MAX_PEN))

        # participation covariates: starter bat side (0 L / 1 R / 2 S),
        # every pitcher's throwing hand, and the catcher flag per slot
        side_code = {"L": 0, "R": 1, "S": 2}
        bat_side = np.array(
            [side_code.get(str(self._bats.get(players[i], "R")), 1)
             for i in range(18)], dtype=np.int8)
        pit_throws = np.array(
            [0 if str(self._throws.get(p, "R")) == "L" else 1
             for p in players], dtype=np.int8)
        slot_is_c = np.array(
            [1 if players[i] in self._catchers else 0
             for i in range(18)], dtype=np.int8)

        adj = self._sim_adjusters(players, season, away, home, date,
                                  pen_rows_away, pen_rows_home)
        self._apply_pen_fatigue(avec, adj["pen_fat_f"])
        lat = dict(self.latent)
        prep = sim.GamePrep(
            n_players=n_players, starters=[18, 19], avec=avec,
            a2vec=a2vec, haz_grid=haz,
            relief_exit=adj["relief_exit"], pen_order=pen_order,
            sb_att=sb_att, sb_suc=sb_suc, sb_state=sb_state,
            patterns=self.patterns,
            latent=lat, bench_rows=bench_rows,
            part_haz=self.part_haz, bat_side=bat_side,
            pit_throws=pit_throws, slot_is_c=slot_is_c,
            run_z=adj["run_z"], dp_z=adj["dp_z"],
            arm_eff=adj["arm_eff"], stretch=adj["stretch"],
            stretch_z=adj["stretch_z"], pen_hi=adj["pen_hi"],
            pen_lo=adj["pen_lo"],
            pre_wp=adj["pre_wp"],
            pen_rank_cum=self._pen_rank_cum,
            pre_pk=self.preevents["pickoff_out_per_pa_runners_on"])
        meta = dict(players=players, names=[self._name(p) if p >= 0
                                            else "League Avg"
                                            for p in players],
                    season=season, away=away, home=home,
                    career_g=[self._career_g.get(p, 0) for p in players],
                    career_gp=[self._career_gp.get(p, 0)
                               for p in players])
        return prep, meta

    def _rest(self, ppid, date):
        # per-player sorted date arrays, built once: the same strictly-
        # before lookup without a 2M-row scan per call
        cache = getattr(self, "_pid_dates", None)
        if cache is None:
            gp = self.stores.raw["gp"]
            cache = {int(p): np.sort(g["Date"].values)
                     for p, g in gp.groupby("PlayerId")}
            self._pid_dates = cache
        try:
            arr = cache.get(int(ppid))
        except (TypeError, ValueError):
            arr = None
        if arr is None or arr.size == 0:
            return np.nan
        d = pd.Timestamp(date)
        i = int(np.searchsorted(arr, np.datetime64(d), side="left"))
        if i == 0:
            return np.nan
        return float((d - pd.Timestamp(arr[i - 1])).days)

    def _pen_np3(self, team, date):
        """Own-pen pitches over the last 3 days (pen fatigue: a gassed
        pen stretches the starter's leash). Mirrors teamctx pen_np3."""
        gp = getattr(self, "_gp_team_cache", {}).get(team)
        if gp is None:
            return np.nan
        d = pd.Timestamp(date)
        rec = gp[(gp.Date >= d - pd.Timedelta(days=3)) & (gp.Date < d)
                 & (pd.to_numeric(gp.GS, errors="coerce") == 0)]
        if not len(rec):
            return 0.0
        return float(pd.to_numeric(rec.NP, errors="coerce")
                     .fillna(0).sum())

    def _league_start_outs(self, season):
        """League mean starter outs for a season (cached); NaN when the
        season isn't (sufficiently) in the game logs."""
        cache = getattr(self, "_lg_outs_cache", None)
        if cache is None:
            cache = self._lg_outs_cache = {}
        if season not in cache:
            gp = self.stores.raw["gp"]
            gs = gp[(pd.to_numeric(gp.GS, errors="coerce") == 1)
                    & (gp.Date.dt.year == season)]
            if len(gs) < 500:
                cache[season] = np.nan
            else:
                ipn = pd.to_numeric(gs.IP, errors="coerce").fillna(0)
                outs = (ipn.astype(int) * 3
                        + np.round((ipn % 1) * 10))
                cache[season] = float(outs.mean())
        return cache[season]

    def _team_hook(self, team, date):
        """As-of team hook tendency (mirrors hazard_table.team_hook):
        season-to-date mean starter outs blended K_HOOK starts with the
        club's prior season (itself shrunk K_HOOK_PREV toward league),
        centered on the prior-season league mean. 0 = average leash."""
        gp = getattr(self, "_gp_team_cache", {}).get(team)
        if gp is None:
            return np.nan
        d = pd.Timestamp(date)
        cache = getattr(self, "_hook_cache", None)
        if cache is None:
            cache = self._hook_cache = {}
        key = (team, str(d.date()))
        if key in cache:
            return cache[key]
        lg = self._league_start_outs(d.year - 1)
        if lg != lg:
            cache[key] = np.nan
            return np.nan

        def _outs_sum(fr):
            ipn = pd.to_numeric(fr.IP, errors="coerce").fillna(0)
            return float((ipn.astype(int) * 3
                          + np.round((ipn % 1) * 10)).sum())

        gs = pd.to_numeric(gp.GS, errors="coerce") == 1
        prv = gp[gs & (gp.Date.dt.year == d.year - 1)]
        cur = gp[gs & (gp.Date >= pd.Timestamp(d.year, 1, 1))
                 & (gp.Date < d)]
        pmean = ((_outs_sum(prv) + F.K_HOOK_PREV * lg)
                 / (len(prv) + F.K_HOOK_PREV))
        th = ((_outs_sum(cur) + F.K_HOOK * pmean)
              / (len(cur) + F.K_HOOK)) - lg
        cache[key] = float(th)
        return cache[key]

    def _hazard_grid(self, ppid, date, season, team=None, post=0.0):
        leash = self.fstores.get("panel_leash")
        lz = None
        if leash is not None:
            mine = leash[(leash.PlayerId == ppid)
                         & (leash.Date < pd.Timestamp(date))]
            if len(mine):
                lz = mine.sort_values("Date").iloc[-1]
        starts = float(lz["starts_d"]) if lz is not None else np.nan
        np_avg = (float(lz["np_sum_d"]) / max(starts, 1e-9)
                  if lz is not None else np.nan)
        bf_avg = (float(lz["bf_sum_d"]) / max(starts, 1e-9)
                  if lz is not None else np.nan)
        ppb = (np_avg / bf_avg) if np_avg and bf_avg and bf_avg > 0 \
            else 3.9
        # start-length dispersion from the leash panel's squared sums
        # (decay factors cancel in the ratio)
        outs_sd = np.nan
        if (lz is not None and "outs2_sum_d" in lz.index
                and starts == starts and starts >= 5):
            mu_o = float(lz["outs_sum_d"]) / max(starts, 1e-9)
            var_o = (float(lz["outs2_sum_d"]) / max(starts, 1e-9)
                     - mu_o ** 2)
            outs_sd = float(np.sqrt(max(var_o, 0.0)))
        # previous-start regime (ramp-up / opener-short) + IL recency
        d = pd.Timestamp(date)
        gp = self.stores.raw["gp"]
        prev = gp[(gp.PlayerId == ppid)
                  & (pd.to_numeric(gp.GS, errors="coerce") == 1)
                  & (gp.Date < d)]
        gap_days, ramp, prev_short = np.nan, 0.0, 0.0
        if len(prev):
            last = prev.sort_values("Date").iloc[-1]
            gap_days = float((d - last.Date).days)
            np_last = pd.to_numeric(last.NP, errors="coerce")
            if pd.notna(np_last):
                ramp = float(np_last < F.RAMP_NP)
            ip_last = pd.to_numeric(last.IP, errors="coerce")
            if pd.notna(ip_last):
                outs_last = int(ip_last) * 3 + round((ip_last % 1) * 10)
                prev_short = float(outs_last <= F.SHORT_START_OUTS)
        il_ret30 = 0.0
        il = self.fstores.get("il_stints")
        if il is not None:
            mine = il[(il.PlayerId == ppid) & (il.Date <= d)]
            if len(mine):
                act = mine.sort_values("Date").iloc[-1]["act_date"]
                il_ret30 = float((d - act).days <= 30)
        bf = np.arange(41)
        runs = np.arange(11)
        B, R = np.meshgrid(bf, runs, indexing="ij")
        rows = pd.DataFrame(dict(
            bf=B.ravel(), cum_pitches=(B * ppb).ravel(),
            tto=(1 + B // 9).clip(1, 4).ravel(),
            inning=(1 + B / 4.3).ravel(), outs=1,
            score_diff=0, k_so_far=(B * 0.22).ravel(),
            br_so_far=(B * 0.30).ravel(), runs_so_far=R.ravel(),
            rest_p=self._rest(ppid, pd.Timestamp(date)),
            leash_np=np_avg, leash_bf=bf_avg, leash_starts=starts,
            season_idx=season - 2015,
            gap_days=gap_days, ramp=ramp, prev_short=prev_short,
            il_ret30=il_ret30, outs_sd=outs_sd,
            pen_np3=(self._pen_np3(team, date) if team else np.nan),
            post=float(post),
            team_hook=(self._team_hook(team, date) if team
                       else np.nan)))
        Xh = rows[self.hz["features"]].astype(np.float32)
        p = self.hz["iso"].predict(
            self.hz["model"].predict_proba(Xh)[:, 1])
        return p.reshape(41, 11)

    def _sb_matrices(self, players, pit_rows, bat_rows_all, date, season,
                     away, home, post=0.0):
        n_players = len(players)
        sprint = self.fstores["sprint"]
        yr = season
        spd = dict(zip(
            sprint.loc[sprint.Year == yr, "PlayerId"],
            sprint.loc[sprint.Year == yr, "SprintSpeed"]))
        sb_cache = getattr(self, "_sb_rate_cache", None)
        if sb_cache is None:
            sb_cache = {}
            self._sb_rate_cache = sb_cache
        if season not in sb_cache:
            ps = pd.read_csv(DATA / "mlb_pitching_stats.csv",
                             encoding="utf-8-sig",
                             usecols=["Year", "PlayerId", "SB", "CS",
                                      "PK", "TBF"])
            ps = ps[pd.to_numeric(ps.Year, errors="coerce")
                    == season - 1]
            for c in ("SB", "CS", "PK", "TBF"):
                ps[c] = pd.to_numeric(ps[c], errors="coerce").fillna(0)
            sb_cache[season] = (
                dict(zip(ps.PlayerId, ps.SB / ps.TBF.clip(lower=50))),
                dict(zip(ps.PlayerId, (ps.CS + ps.PK)
                         / (ps.SB + ps.CS + ps.PK).clip(lower=3))))
        sbr, csr = sb_cache[season]
        deft = self.fstores["defense_team"]
        drow = {t: deft[(deft.Team == t) & (deft.Year == season)]
                for t in (away, home)}

        def team_pop(t):
            r = drow[t]
            return (float(r.PopTime.iloc[0]) if len(r) and
                    pd.notna(r.PopTime.iloc[0]) else np.nan)

        era_new = float(season >= 2023)
        att = np.zeros((n_players, n_players))
        suc = np.zeros((n_players, n_players))
        shift = self.sb["success_logit_shift"][
            "post2023" if era_new else "pre2023"]
        scale = self.sb["scale"][
            "post2023" if era_new else "pre2023"]["attempt_scale"]
        spc = self.sb.get("speed_center", 27.3)
        rows_a, rows_s, pos = [], [], []
        for brow in bat_rows_all:
            bpid = players[brow]
            for prow in pit_rows:
                ppid = players[prow]
                # the catcher backing this pitcher belongs to his club
                fld = away if (prow == 18 or
                               20 <= prow < 20 + MAX_PEN) else home
                lhp = float(str(self._throws.get(ppid, "R")) == "L")
                sspd = (spd.get(bpid, np.nan) if bpid >= 0 else np.nan)
                rows_a.append(dict(
                    SprintSpeed=sspd,
                    speed_miss=float(np.isnan(sspd)),
                    speed2=(sspd - spc) ** 2,
                    sb_allowed_rate=sbr.get(ppid, np.nan),
                    cs_rate=csr.get(ppid, np.nan),
                    PopTime=team_pop(fld), CSAA=np.nan,
                    outs=1, outs1=1.0, score_close=1.0,
                    era_new=era_new, lhp=lhp, post=post))
                rows_s.append(dict(
                    SprintSpeed=spd.get(bpid, np.nan),
                    cs_rate=csr.get(ppid, np.nan),
                    PopTime=team_pop(fld), CSAA=np.nan, lhp=lhp,
                    era_new=era_new, post=post))
                pos.append((brow, prow))
        A = pd.DataFrame(rows_a)[self.sb["att_features"]]
        S_ = pd.DataFrame(rows_s)[self.sb["suc_features"]]
        pa_ = self.sb["attempt"].predict_proba(A)[:, 1] * scale
        p_raw = self.sb["success"].predict_proba(S_)[:, 1]
        logit = np.log(np.clip(p_raw, 1e-6, 1 - 1e-6)
                       / np.clip(1 - p_raw, 1e-6, 1))
        ps_ = 1 / (1 + np.exp(-(logit + shift)))
        for (brow, prow), a_, s_ in zip(pos, pa_, ps_):
            att[brow, prow] = a_
            suc[brow, prow] = s_
        return att, suc

    def _sb_state(self, season):
        """Sim-time steal-state vector [outs0, outs2, sc_far, scale]:
        raw logit deltas off the (outs=1, close-game) serve baseline
        plus the era attempt scale — None for pre-state artifacts."""
        st = self.sb.get("att_state_logit")
        if not st:
            return None
        era = "post2023" if season >= 2023 else "pre2023"
        return np.array([st["outs0"], st["outs2"], st["sc_far"],
                         self.sb["scale"][era]["attempt_scale"]],
                        dtype=np.float32)

    # ------------------------------------------------------ slate run

    def _game_effects(self, spec):
        """Venue HR factor + ump run factor for the residual heads'
        game-grain context (1.0 when unknown)."""
        season = int(spec.get("season")
                     or pd.Timestamp(spec["date"]).year)
        park_hr = ump_r = 1.0
        pf = self.fstores.get("park_factors")
        if pf is not None:
            m = pf[(pf.Venue == (spec.get("venue") or ""))
                   & (pf.Year == season)]
            if len(m):
                park_hr = float(m.pf_HR.iloc[0])
        uf = self.fstores.get("ump_factors")
        ump = spec.get("hp_ump_id")
        if uf is not None and ump:
            m = uf[(uf.HpUmpId == int(ump)) & (uf.Year == season)]
            if len(m):
                ump_r = float(m.uf_R.iloc[0])
        return park_hr, ump_r

    def predict_slate(self, specs, n_sims=N_SIMS, progress=None):
        tick = progress or (lambda m: None)
        hctx = None
        tick("building slate context...")
        try:
            hctx = F.slate_context(specs)
            for sp, cx in zip(specs, hctx):
                cx["park_hr"], cx["ump_r"] = self._game_effects(sp)
                sp["_ctx"] = cx     # Elo into the A1 feature rows
        except Exception as e:              # noqa: BLE001
            print(f"slate context unavailable ({e}); serving without "
                  f"Elo/heads context", flush=True)
        out = []
        for gi, spec in enumerate(specs):
            tick(f"game {gi + 1}/{len(specs)}: preparing...")
            prep, meta = self.prepare_game(spec, n_sims=n_sims)
            tick(f"game {gi + 1}/{len(specs)}: simulating...")
            res = sim.run(prep, n_sims=n_sims,
                          seed=int(pd.Timestamp(spec["date"]).toordinal())
                          * 100 + gi,
                          season=meta["season"],
                          is_dh_game=bool(spec.get("is_dh")))
            res["meta"] = meta
            res["spec"] = spec
            res["calib"] = self.calib
            if hctx is not None:
                res["heads"] = self.heads
                res["hctx"] = hctx[gi]
                res["hdef"] = self._defense
            out.append(res)
        return out


# -------------------------------------------------------- aggregation

# market families for the cross-line-coherent output calibrators: ONE
# shared monotone map per family (fit by evaluate.py --fit-calibrators),
# so calibration can never break P(2+ hits) <= P(1+ hit)
COL_FAM = {"HR": "hr", "Hit": "h", "2+ Hits": "h", "Single": "b1",
           "Double": "b2", "Triple": "b3", "2+ TB": "tb", "3+ TB": "tb",
           "4+ TB": "tb", "Run": "r", "2+ Runs": "r", "RBI": "rbi",
           "2+ RBI": "rbi", "H+R+RBI 2+": "hrr", "H+R+RBI 3+": "hrr",
           "H+R+RBI 4+": "hrr", "BB": "bb", "SB": "sb", "K": "bk",
           "2+ K": "bk", "3+ K": "bk"}
PIT_FAM = {"K > ": "pk", "Outs > ": "pout", "Hits > ": "pha",
           "BB > ": "pbb", "ER > ": "per"}
MKT_FAM = {"batter_hits": "h", "batter_home_runs": "hr",
           "batter_total_bases": "tb", "batter_runs_scored": "r",
           "batter_rbis": "rbi", "batter_walks": "bb",
           "batter_stolen_bases": "sb", "batter_singles": "b1",
           "batter_doubles": "b2", "batter_hits_runs_rbis": "hrr",
           "pitcher_strikeouts": "pk", "pitcher_outs": "pout",
           "pitcher_hits_allowed": "pha", "pitcher_walks": "pbb",
           "pitcher_earned_runs": "per", "h2h": "ml", "totals": "tot",
           "team_totals": "tt"}


def _cal(calib, fam, p):
    """Apply a family's shared monotone calibrator (identity when the
    family has none fitted). Output is clamped strictly inside (0, 1):
    no calibrator may ever emit a hard 0/1 into grading or EV math."""
    cal = (calib or {}).get(fam)
    if cal is None or p is None:
        return p
    return float(np.clip(cal.predict(np.array([p]))[0],
                         1e-6, 1 - 1e-6))


def _dec(american):
    """American odds -> decimal payout multiplier (stake included)."""
    try:
        a = float(american)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(a) or a == 0:
        return None
    return 1 + (a / 100.0 if a > 0 else 100.0 / -a)


def _line_of(market):
    """Numeric line of a 'X > n.5'-style market name (NaN otherwise) —
    shared with heads.py so train and serve tokenize identically."""
    try:
        return float(str(market).split(">")[-1])
    except (TypeError, ValueError):
        return np.nan


def _apply_heads(heads, ctx, deft, season, bat_rows, pit_rows, grow,
                 away, home):
    """Residual-head corrections, applied AFTER the family calibrators:
    per market family, a shallow-GBM logit adjustment from team context
    (heads.py --train). Feature construction must mirror
    heads._features exactly — the saved feature list orders the frame,
    so a missing column fails loudly rather than silently misaligning.
    "ml" is never adjusted here (min-rows skipped at train time; its
    trained market string "home ML" has no workbook column)."""
    jobs = {}

    def add(cont, col, fam, home_flag):
        h = heads.get(fam)
        p = cont.get(col)
        if h is None or p is None:
            return
        jobs.setdefault(fam, []).append((cont, col, home_flag, float(p)))

    for br in bat_rows:
        hf = int(br["Team"] == home)
        for col, fam in COL_FAM.items():
            add(br, col, fam, hf)
    for pr_ in pit_rows:
        hf = int(pr_["Team"] == home)
        for col in list(pr_):
            for pref, fam in PIT_FAM.items():
                if str(col).startswith(pref):
                    add(pr_, col, fam, hf)
    for x in TEAM_TOTAL_LINES:
        add(grow, f"Away Runs > {x}", "tt", 0)
        add(grow, f"Home Runs > {x}", "tt", 1)
    for x in TOTAL_LINES:
        add(grow, f"Runs > {x}", "tot", 1)

    for fam, items in jobs.items():
        h = heads[fam]
        feats = []
        for _cont, col, hf, p in items:
            own, opp = ("home", "away") if hf else ("away", "home")
            own_t, opp_t = (home, away) if hf else (away, home)
            f = {}
            for c in HEAD_CTX:
                f[f"own_{c}"] = ctx.get(f"{own}_{c}", np.nan)
                f[f"opp_{c}"] = ctx.get(f"{opp}_{c}", np.nan)
            f["elo_diff"] = ctx.get("elo_diff", np.nan)
            f["own_def_oaa"], f["own_frame"] = deft.get(
                (season, own_t), (np.nan, np.nan))
            f["opp_def_oaa"], f["opp_frame"] = deft.get(
                (season, opp_t), (np.nan, np.nan))
            f["home"] = float(hf)
            f["line"] = _line_of(col)
            f["p_dist"] = abs(p - 0.5)
            f["park_hr"] = ctx.get("park_hr", 1.0)
            f["ump_r"] = ctx.get("ump_r", 1.0)
            feats.append(f)
        X = pd.DataFrame(feats)[h["features"]]
        raw = h["bst"].predict(X, num_iteration=h["best_iter"],
                               raw_score=True)
        for (cont, col, hf, p), r in zip(items, raw):
            pc = min(max(p, 1e-6), 1 - 1e-6)
            z = np.log(pc / (1 - pc)) + float(r)
            cont[col] = float(np.clip(1.0 / (1.0 + np.exp(-z)),
                                      1e-6, 1 - 1e-6))

    # cross-line coherence guard: within one row's ladder (same
    # container, family, side) the workbook authors columns in
    # ascending line/count order, so probabilities must be
    # non-increasing — per-line head adjustments may not cross them
    # (the family calibrators are shared monotone maps for exactly
    # this reason; the heads must not undo that invariant)
    ladders = {}
    for fam, items in jobs.items():
        for cont, col, hf, _p in items:
            ladders.setdefault((id(cont), fam, hf),
                               []).append((cont, col))
    for lad in ladders.values():
        lo = None
        for cont, col in lad:
            v = cont[col]
            if lo is not None and v > lo:
                v = lo
                cont[col] = v
            lo = v


def game_frame(res):
    """One sim result -> the workbook rows for each sheet, with the
    family output calibrators applied to every probability."""
    t = res["tensor"]
    s = sim.SIDX
    calib = res.get("calib") or {}
    meta, spec = res["meta"], res["spec"]
    away, home = meta["away"], meta["home"]
    label = f"{away}@{home}"
    gnum = 2 if spec.get("dh_game2") else 1

    bat_rows = []
    for brow in range(18):
        pid = meta["players"][brow]
        if pid < 0:
            continue
        sub = t[:, brow, :]
        tb = (sub[:, s["B1"]] + 2 * sub[:, s["B2"]]
              + 3 * sub[:, s["B3"]] + 4 * sub[:, s["HR"]])
        hrr = sub[:, s["H"]] + sub[:, s["R"]] + sub[:, s["RBI"]]
        row = {
            "Game": label,
            "Team": away if brow < 9 else home,
            "G#": gnum,
            "Slot": brow % 9 + 1,
            "Name": meta["names"][brow],
            "ID": pid,
            "Career G": meta["career_g"][brow],
            "HR": float((sub[:, s["HR"]] >= 1).mean()),
            "xTB": round(float(tb.mean()), 2),
            "2+ TB": _tail_prob(tb, 1.5),
            "3+ TB": _tail_prob(tb, 2.5),
            "4+ TB": _tail_prob(tb, 3.5),
            "xH": round(float(sub[:, s["H"]].mean()), 2),
            "Hit": float((sub[:, s["H"]] >= 1).mean()),
            "2+ Hits": _tail_prob(sub[:, s["H"]], 1.5),
            "xRBI": round(float(sub[:, s["RBI"]].mean()), 2),
            "RBI": float((sub[:, s["RBI"]] >= 1).mean()),
            "2+ RBI": _tail_prob(sub[:, s["RBI"]], 1.5),
            "xR": round(float(sub[:, s["R"]].mean()), 2),
            "Run": float((sub[:, s["R"]] >= 1).mean()),
            "2+ Runs": _tail_prob(sub[:, s["R"]], 1.5),
            "xHRR": round(float(hrr.mean()), 2),
            "H+R+RBI 2+": _tail_prob(hrr, 1.5),
            "H+R+RBI 3+": _tail_prob(hrr, 2.5),
            "H+R+RBI 4+": _tail_prob(hrr, 3.5),
            "SB": float((sub[:, s["SB"]] >= 1).mean()),
            "xSO": round(float(sub[:, s["K"]].mean()), 2),
            "K": float((sub[:, s["K"]] >= 1).mean()),
            "2+ K": _tail_prob(sub[:, s["K"]], 1.5),
            "3+ K": _tail_prob(sub[:, s["K"]], 2.5),
            "Single": float((sub[:, s["B1"]] >= 1).mean()),
            "Double": float((sub[:, s["B2"]] >= 1).mean()),
            "Triple": float((sub[:, s["B3"]] >= 1).mean()),
            "xBB": round(float(sub[:, s["BB"]].mean()), 2),
            "BB": float((sub[:, s["BB"]] >= 1).mean()),
        }
        for c, fam in COL_FAM.items():
            row[c] = _cal(calib, fam, row[c])
        bat_rows.append(row)

    pit_rows = []
    for prow, team, opp in ((18, away, home), (19, home, away)):
        pid = meta["players"][prow]
        if pid < 0:
            continue
        sub = t[:, prow, :]
        row = {"Game": label, "Team": team, "Opponent": opp, "G#": gnum,
               "Name": meta["names"][prow], "ID": pid,
               "xK": round(float(sub[:, s["PK_"]].mean()), 2)}
        for x in K_LINES:
            row[f"K > {x}"] = _tail_prob(sub[:, s["PK_"]], x)
        row["xER"] = round(float(sub[:, s["PER"]].mean()), 2)
        for x in PER_LINES:
            row[f"ER > {x}"] = _tail_prob(sub[:, s["PER"]], x)
        row["xOuts"] = round(float(sub[:, s["OUTS"]].mean()), 2)
        for x in OUT_LINES:
            row[f"Outs > {x}"] = _tail_prob(sub[:, s["OUTS"]], x)
        row["xHits"] = round(float(sub[:, s["PH"]].mean()), 2)
        for x in PHA_LINES:
            row[f"Hits > {x}"] = _tail_prob(sub[:, s["PH"]], x)
        row["xBB"] = round(float(sub[:, s["PBB"]].mean()), 2)
        for x in PBB_LINES:
            row[f"BB > {x}"] = _tail_prob(sub[:, s["PBB"]], x)
        for c in row:
            for pref, fam in PIT_FAM.items():
                if c.startswith(pref):
                    row[c] = _cal(calib, fam, row[c])
        pit_rows.append(row)

    sc = res["score"]
    total = sc.sum(axis=1)
    home_wp = _cal(calib, "ml",
                   float((sc[:, 1] > sc[:, 0]).mean()))
    lineup_hr = float(t[:, :18, s["HR"]].sum(axis=1).mean())
    grow = {"Game": label, "Date": str(spec.get("date", "")),
            "Venue": spec.get("venue", ""),
            "Winner": home if home_wp >= 0.5 else away,
            "Win Prob": max(home_wp, 1 - home_wp)}
    for x in TEAM_TOTAL_LINES:
        grow[f"Away Runs > {x}"] = _cal(calib, "tt",
                                        float((sc[:, 0] > x).mean()))
    for x in TEAM_TOTAL_LINES:
        grow[f"Home Runs > {x}"] = _cal(calib, "tt",
                                        float((sc[:, 1] > x).mean()))
    grow["Away Score"] = round(float(sc[:, 0].mean()), 2)
    grow["Home Score"] = round(float(sc[:, 1].mean()), 2)
    grow["Total Runs"] = round(float(total.mean()), 2)
    for x in TOTAL_LINES:
        grow[f"Runs > {x}"] = _cal(calib, "tot",
                                   float((total > x).mean()))
    grow["Lineup HRs"] = round(lineup_hr, 2)
    grow["_home_wp"] = home_wp
    grow["_nrfi"] = float((res["runs_i1"].sum(axis=1) == 0).mean())
    if res.get("heads") and res.get("hctx") is not None:
        _apply_heads(res["heads"], res["hctx"], res.get("hdef") or {},
                     int(meta["season"]), bat_rows, pit_rows, grow,
                     away, home)
    return dict(bat=bat_rows, pit=pit_rows, game=grow)


GAME_COLS = (["Game", "Date", "Venue", "Winner", "Win Prob",
              "Away Score"]
             + [f"Away Runs > {x}" for x in TEAM_TOTAL_LINES]
             + ["Home Score"]
             + [f"Home Runs > {x}" for x in TEAM_TOTAL_LINES]
             + ["Total Runs"]
             + [f"Runs > {x}" for x in TOTAL_LINES]
             + ["Lineup HRs"])

_BAT_STAT = {"batter_hits": "H", "batter_home_runs": "HR",
             "batter_runs_scored": "R", "batter_rbis": "RBI",
             "batter_walks": "BB", "batter_stolen_bases": "SB",
             "batter_singles": "B1", "batter_doubles": "B2"}
_PIT_STAT = {"pitcher_strikeouts": "PK_", "pitcher_outs": "OUTS",
             "pitcher_hits_allowed": "PH", "pitcher_walks": "PBB",
             "pitcher_earned_runs": "PER"}
_PIT_WORD = {"pitcher_strikeouts": "strikeouts", "pitcher_outs": "outs",
             "pitcher_hits_allowed": "hits allowed",
             "pitcher_walks": "walks", "pitcher_earned_runs":
             "earned runs"}
def _prop_name(market, line):
    n = int(line + 0.5)
    if market == "batter_hits":
        return "1+ hit" if n == 1 else f"{n}+ hits"
    if market == "batter_home_runs":
        return "1+ HR" if n == 1 else f"{n}+ HR"
    if market == "batter_total_bases":
        return f"{n}+ total bases"
    if market == "batter_runs_scored":
        return "run scored" if n == 1 else f"{n}+ runs scored"
    if market == "batter_rbis":
        return f"{n}+ RBI"
    if market == "batter_hits_runs_rbis":
        return f"{n}+ H+R+RBI"
    if market == "batter_walks":
        return "1+ walk" if n == 1 else f"{n}+ walks"
    if market == "batter_stolen_bases":
        return "stolen base"
    if market == "batter_singles":
        return "1+ single" if n == 1 else f"{n}+ singles"
    if market == "batter_doubles":
        return "1+ double" if n == 1 else f"{n}+ doubles"
    if market in _PIT_WORD:
        return f"pitcher {_PIT_WORD[market]} o{line}"
    return market


def _best_price(group, col):
    """Most generous posted price for one side across books."""
    best, book = None, ""
    for _, r in group.iterrows():
        d = _dec(r.get(col))
        if d is None:
            continue
        if best is None or d > best[0]:
            best = (d, r.get(col))
            book = r.get("Book", "")
    if best is None:
        return None, None, ""
    return best[0], int(float(best[1])), book


def build_bets(out, date):
    """Model vs the day's captured odds: every side clearing MIN_EV at
    the best posted price, sorted by EV descending."""
    store = Path(O.DEFAULT_STORE)
    if not store.exists():
        return []
    try:
        odds = pd.read_csv(store, encoding="utf-8-sig", low_memory=False)
    except Exception:                                # noqa: BLE001
        return []
    odds = odds[odds.Date.astype(str) == str(date)]
    if odds.empty:
        return []

    s = sim.SIDX
    by_pid, by_game = {}, {}
    for res in out:
        meta = res["meta"]
        label = f'{meta["away"]}@{meta["home"]}'
        by_game[meta["home"]] = (res, label)
        for row_i, pid in enumerate(meta["players"]):
            if pid >= 0 and row_i < 20:
                by_pid.setdefault(int(pid), []).append((res, row_i, label))

    rows = []

    def stat_counts(res, row_i, market):
        t = res["tensor"]
        if market == "batter_total_bases":
            return (t[:, row_i, s["B1"]] + 2 * t[:, row_i, s["B2"]]
                    + 3 * t[:, row_i, s["B3"]]
                    + 4 * t[:, row_i, s["HR"]])
        if market == "batter_hits_runs_rbis":
            return (t[:, row_i, s["H"]] + t[:, row_i, s["R"]]
                    + t[:, row_i, s["RBI"]])
        if market in _BAT_STAT:
            return t[:, row_i, s[_BAT_STAT[market]]]
        if market in _PIT_STAT:
            return t[:, row_i, s[_PIT_STAT[market]]]
        return None

    def emit(game, player, team, prop, side, line, p_model, fair_side,
             group, price_col, note_bits):
        # a de-vigged two-sided market is required: without it a stray
        # longshot price has no sanity anchor and EV explodes
        if fair_side is None or p_model is None:
            return
        dec_best, amer, book = _best_price(group, price_col)
        if dec_best is None:
            return
        ev = p_model * dec_best - 1
        if ev < MIN_EV:
            return
        books = group["Book"].nunique()
        bits = list(note_bits)
        if books == 1:
            bits.append("1 book")
        rows.append({
            "Game": game, "Player": player, "Team": team, "Prop": prop,
            "Side": side, "Line": line, "Model %": p_model,
            "Mkt %": fair_side, "Edge": p_model - fair_side,
            "Best Odds": amer, "Book": book, "EV%": ev, "Books": books,
            "Note": "; ".join(bits) if bits else None})

    # player props
    pl = odds[pd.to_numeric(odds.PlayerId, errors="coerce").notna()]
    for (pid, market, line), g in pl.groupby(
            [pd.to_numeric(pl.PlayerId, errors="coerce").astype("int64"),
             "Market", "Line"]):
        try:
            line = float(line)
        except (TypeError, ValueError):
            continue
        hits = by_pid.get(int(pid))
        if not hits:
            continue
        res, row_i, label = hits[0]
        counts = stat_counts(res, row_i, market)
        if counts is None:
            continue
        p_over = _cal(res.get("calib"), MKT_FAM.get(market),
                      _tail_prob(counts, line))
        fair = O.sharp_fair(g.to_dict("records"))
        meta = res["meta"]
        name = meta["names"][row_i]
        team = (meta["away"] if (row_i < 9 or row_i == 18)
                else meta["home"])
        career = (meta["career_g"][row_i] if row_i < 18
                  else meta["career_gp"][row_i])
        note = ["rookie <50 G"] if career < ROOKIE_G else []
        prop = _prop_name(market, line)
        emit(label, name, team, prop, "Over", line, p_over,
             fair, g, "OverPrice", note)
        emit(label, name, team, prop, "Under", line,
             1 - p_over, (1 - fair) if fair is not None else None, g,
             "UnderPrice", note)

    # game markets: h2h (Over = home side) and totals
    gm = odds[odds.Market.isin(["h2h", "totals"])
              & pd.to_numeric(odds.PlayerId, errors="coerce").isna()]
    for (team, market, line), g in gm.groupby(["Team", "Market", "Line"],
                                              dropna=False):
        hit = by_game.get(team)
        if hit is None:
            continue
        res, label = hit
        meta = res["meta"]
        fair = O.sharp_fair(g.to_dict("records"))
        if market == "h2h":
            hw = _cal(res.get("calib"), "ml",
                      float((res["score"][:, 1]
                             > res["score"][:, 0]).mean()))
            note = ["winner: no proven edge vs. always-home"]
            emit(label, "", meta["home"], "moneyline", meta["home"], "",
                 hw, fair, g, "OverPrice", note)
            emit(label, "", meta["away"], "moneyline", meta["away"], "",
                 1 - hw, (1 - fair) if fair is not None else None, g,
                 "UnderPrice", note)
        else:
            try:
                line = float(line)
            except (TypeError, ValueError):
                continue
            total = res["score"].sum(axis=1)
            p_over = _cal(res.get("calib"), "tot",
                          float((total > line).mean()))
            emit(label, "", "", "total runs", "Over", line, p_over,
                 fair, g, "OverPrice", [])
            emit(label, "", "", "total runs", "Under", line, 1 - p_over,
                 (1 - fair) if fair is not None else None, g,
                 "UnderPrice", [])

    rows.sort(key=lambda r: -(r["EV%"] or 0))
    return rows


def save_excel_slate(specs, out, path=None):
    """Aggregate sim results into the workbook (Batter Props, Pitching
    Props, Games, Bets). `out` is predict_slate()'s list."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, \
        Side
    from openpyxl.utils import get_column_letter

    frames = [game_frame(r) for r in out]
    date = str(specs[0]["date"]) if specs else dt.date.today().isoformat()
    PRED_DIR.mkdir(exist_ok=True)
    if path is None:
        path = PRED_DIR / f"{date}.xlsx"
        k = 2
        while Path(path).exists():
            path = PRED_DIR / f"{date}_{k}.xlsx"
            k += 1

    bet_rows = build_bets(out, date)

    wb = Workbook()
    white_bold = Font(color="FFFFFFFF", bold=True)
    red_font = Font(color=ROOKIE_RED)
    row_tint = PatternFill("solid", fgColor=BETS_ROW_TINT)
    thin = Side(style="thin", color="FFB7B7B7")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    def write_sheet(ws, cols, rows, pct_cols, header_color=NAVY,
                    all_row_fill=None, width_cap=40):
        head_fill = PatternFill("solid", fgColor=header_color)
        widths = [len(str(c)) for c in cols]
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=j, value=c)
            cell.fill = head_fill
            cell.font = white_bold
            cell.border = box
            cell.alignment = center
        for i, r in enumerate(rows, start=2):
            for j, c in enumerate(cols, start=1):
                v = r.get(c)
                cell = ws.cell(row=i, column=j, value=v)
                cell.border = box
                cell.alignment = center
                if all_row_fill is not None:
                    cell.fill = all_row_fill
                shown = "" if v is None else str(v)
                if c in pct_cols and isinstance(v, float):
                    cell.number_format = "0.0%"
                    shown = f"{v:.1%}"
                elif c == "Best Odds" and v is not None:
                    cell.number_format = "+0;-0"
                    shown = f"{int(v):+d}"
                elif isinstance(v, float):
                    shown = f"{v:.2f}"
                if c == "Career G" and isinstance(v, (int, float)) \
                        and v < ROOKIE_G:
                    cell.font = red_font
                widths[j - 1] = max(widths[j - 1], len(shown))
        ws.freeze_panes = "A2"
        # header dropdown arrows: sort/filter on every column
        ws.auto_filter.ref = (f"A1:{get_column_letter(len(cols))}"
                              f"{len(rows) + 1}")
        for j in range(1, len(cols) + 1):
            # +4 keeps content clear of the filter arrow
            ws.column_dimensions[get_column_letter(j)].width = \
                min(widths[j - 1] + 4, width_cap)

    ws = wb.active
    ws.title = "Batter Props"
    bat_rows = [r for f in frames for r in f["bat"]]
    # board default: most likely homer at the top
    bat_rows.sort(key=lambda r: -(r.get("HR") or 0.0))
    bat_cols = list(bat_rows[0].keys()) if bat_rows else []
    bat_pct = {c for c in bat_cols
               if c not in ("Game", "Team", "G#", "Slot", "Name", "ID",
                            "Career G") and not c.startswith("x")}
    write_sheet(ws, bat_cols, bat_rows, bat_pct)

    ws = wb.create_sheet("Pitching Props")
    pit_rows = [r for f in frames for r in f["pit"]]
    pit_cols = list(pit_rows[0].keys()) if pit_rows else []
    pit_pct = {c for c in pit_cols if ">" in c}
    write_sheet(ws, pit_cols, pit_rows, pit_pct)

    ws = wb.create_sheet("Games")
    game_rows = [f["game"] for f in frames]
    game_pct = {c for c in GAME_COLS if ">" in c or c == "Win Prob"}
    write_sheet(ws, GAME_COLS, game_rows, game_pct)

    ws = wb.create_sheet("Bets")
    bet_cols = ["Game", "Player", "Team", "Prop", "Side", "Line",
                "Model %", "Mkt %", "Edge", "Best Odds", "Book", "EV%",
                "Books", "Note"]
    write_sheet(ws, bet_cols, bet_rows,
                {"Model %", "Mkt %", "Edge", "EV%"},
                header_color=BETS_HEADER, all_row_fill=row_tint)
    if not bet_rows:
        cell = ws.cell(row=2, column=1,
                       value="no captured odds for this date — run "
                             "\"Tools/2) Scrape Odds.py\" near game time")
        cell.border = box
        cell.alignment = center

    wb.save(path)
    return str(path)


# --------------------------------------------------------------- CLI

def main():
    """Headless serve: Data/todays_games.json -> workbook + sims npz.

    The scheduler's serve entry point — the exact equivalent of the
    GUI's Predict button (same spec assembly the GUI does: games sorted
    by start_et, lineups tuple-ified), so a scheduled serve and a GUI
    serve are indistinguishable downstream. Each game's npz product tag
    (projected vs confirmed) comes from the per-side lineup provenance
    the slate scraper recorded; serving early in the day simply yields
    projected-product games, and a later re-serve yields confirmed
    ones.

    Usage:
        python Model/predict.py --serve             # full slate, 20k sims
        python Model/predict.py --serve --sims 4000 # faster smoke test
        python Model/predict.py --serve --json Data/slates/a.json \
            Data/slates/b.json   # batch: one workbook per slate file
    """
    import argparse
    ap = argparse.ArgumentParser(description=main.__doc__)
    ap.add_argument("--serve", action="store_true",
                    help="serve the slate headlessly (GUI-equivalent)")
    ap.add_argument("--sims", type=int, default=N_SIMS)
    ap.add_argument("--json", nargs="+",
                    default=[str(DATA / "todays_games.json")],
                    help="one or more slate JSONs; each gets its own "
                         "workbook (named from its own date)")
    args = ap.parse_args()
    if not args.serve:
        ap.error("nothing to do: pass --serve")

    P = None  # models load once, first slate that has games
    served = 0
    for jpath in args.json:
        payload = json.loads(Path(jpath).read_text(encoding="utf-8"))
        games = payload.get("games") or []
        if not games:
            print(f"{jpath}: no games in the slate file; skipping")
            continue
        specs = sorted(games, key=lambda g: (g.get("start_et") or "99:99",
                                             g.get("away_team") or ""))
        for s in specs:
            for side in ("away", "home"):
                s[f"{side}_lineup"] = [tuple(x) for x in
                                       (s.get(f"{side}_lineup") or [])]

        if P is None:
            P = Predictor(progress=lambda m: print(m, flush=True))
        out = P.predict_slate(specs, n_sims=args.sims,
                              progress=lambda m: print(m, flush=True))
        path = save_excel_slate(specs, out)
        n_conf = sum(1 for s in specs
                     if s.get("away_lineup_src", "mlb") == "mlb"
                     and s.get("home_lineup_src", "mlb") == "mlb")
        print(f"served {len(specs)} games at {args.sims} sims "
              f"({n_conf} confirmed-lineup, {len(specs) - n_conf} "
              f"projected) -> {path}")
        served += 1
    if len(args.json) > 1:
        print(f"batch complete: {served}/{len(args.json)} slates served")


if __name__ == "__main__":
    main()
